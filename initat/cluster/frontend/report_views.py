# Copyright (C) 2016 Gregor Kaufmann, Andreas Lang-Nevyjel
#
# Send feedback to: <g.kaufmann@init.at>, <lang-nevyjel@init.at>
#
# This file is part of icsw-server
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License Version 2 as
# published by the Free Software Foundation.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  USA
#

""" report views """

import base64
import datetime
import json
import logging
import os
import tempfile
import Queue
from io import BytesIO
from threading import Thread

from PIL import Image as PILImage
from PollyReports import Element, Rule, Band
from PollyReports import Image as PollyReportsImage
from PollyReports import Report as PollyReportsReport
from PyPDF2 import PdfFileWriter, PdfFileReader
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.generic import View
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from reportlab.graphics.shapes import Drawing, Rect
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.pagesizes import landscape, letter, A4, A3
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, inch
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, Paragraph, Image

from initat.cluster.backbone.models import device, device_variable, AssetType, PackageTypeEnum, RunStatus, RunResult
from initat.cluster.backbone.models import network
from initat.cluster.backbone.models import user
from initat.cluster.backbone.models.asset import ASSET_DATETIMEFORMAT
from initat.cluster.backbone.models.report import ReportHistory
from initat.cluster.backbone.models.user import AC_READONLY, AC_MODIFY, AC_CREATE, AC_FULL
from initat.cluster.frontend.helper_functions import xml_wrapper

_ = {landscape, letter, A4, A3}

if settings.DEBUG:
    _file_root = os.path.join(settings.FILE_ROOT, "frontend", "static")
    NOCTUA_LOGO_PATH = os.path.join(_file_root, "images", "product", "noctua-flat-trans.png")
else:
    _file_root = settings.ICSW_PROD_WEB_DIR
    NOCTUA_LOGO_PATH = os.path.join(settings.STATIC_ROOT, "noctua-flat-trans.png")

# this makes problems when running package-knife builds on devices without installed icsw-server
# package because DEBUG=False


pdfmetrics.registerFont(TTFont('SourceSansPro-Black',
                               os.path.join(_file_root, "fonts", "SourceSansPro-Black.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-BlackIt',
                               os.path.join(_file_root, "fonts", "SourceSansPro-BlackIt.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-Bold',
                               os.path.join(_file_root, "fonts", "SourceSansPro-Bold.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-BoldIt',
                               os.path.join(_file_root, "fonts", "SourceSansPro-BoldIt.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-ExtraLight',
                               os.path.join(_file_root, "fonts", "SourceSansPro-ExtraLight.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-ExtraLightIt',
                               os.path.join(_file_root, "fonts",
                                            "SourceSansPro-ExtraLightIt.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-It',
                               os.path.join(_file_root, "fonts", "SourceSansPro-It.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-Light',
                               os.path.join(_file_root, "fonts", "SourceSansPro-Light.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-LightIt',
                               os.path.join(_file_root, "fonts", "SourceSansPro-LightIt.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-Regular',
                               os.path.join(_file_root, "fonts", "SourceSansPro-Regular.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-Semibold',
                               os.path.join(_file_root, "fonts", "SourceSansPro-Semibold.ttf")))
pdfmetrics.registerFont(TTFont('SourceSansPro-SemiboldIt',
                               os.path.join(_file_root, "fonts", "SourceSansPro-SemiboldIt.ttf")))


logger = logging.getLogger(__name__)


########################################################################################################################
# Report Generator Classes / Functions
########################################################################################################################

class RowCollector(object):
    def __init__(self):
        self.rows_dict = []
        self.row_info = []
        self.current_asset_type = None

    def reset(self):
        self.rows_dict = []
        self.row_info = []
        self.current_asset_type = None

    def collect(self, _row):
        self.row_info = _row[0:8]

        if self.current_asset_type == AssetType.UPDATE:
            update_name = str(_row[8])
            # update_version = str(_row[9])
            # update_release = str(_row[10])
            # update_kb_idx = str(_row[11])
            install_date = _row[12]
            # update_install_date = str(_row[12])
            update_status = str(_row[13])
            # update_installed = str(_row[15])

            if type(install_date) != str:
                install_date = install_date.strftime(ASSET_DATETIMEFORMAT)

            o = {
                'update_name': update_name,
                'install_date': install_date,
                'update_status': update_status,
            }

            self.rows_dict.append(o)

        elif self.current_asset_type == AssetType.LICENSE:
            license_name = str(_row[8])
            license_key = str(_row[9])

            o = {
                'license_name': license_name,
                'license_key': license_key
            }

            self.rows_dict.append(o)

        elif self.current_asset_type == AssetType.PENDING_UPDATE:
            update_name = str(_row[8])
            # update_version = str(_row[9])
            # update_release = str(_row[10])
            # update_kb_idx = str(_row[11])
            # update_install_date = str(_row[12])
            # update_status = str(_row[13])
            update_optional = str(_row[14])
            # update_installed = str(_row[15])
            update_new_version = str(_row[16])

            o = {
                'update_name': update_name,
                'update_version': update_new_version if update_new_version else "N/A",
                'update_optional': update_optional
            }

            self.rows_dict.append(o)

        elif self.current_asset_type == AssetType.PROCESS:
            process_name = str(_row[8])
            process_id = str(_row[9])

            o = {
                'process_name': process_name,
                'process_id': process_id
            }

            self.rows_dict.append(o)

        elif self.current_asset_type == AssetType.HARDWARE:
            hardware_node_type = str(_row[8])
            hardware_depth = str(_row[9])
            hardware_attributes = str(_row[10])

            o = {
                'hardware_node_type': hardware_node_type,
                'hardware_depth': hardware_depth,
                'hardware_attributes': hardware_attributes,
            }

            self.rows_dict.append(o)

        elif self.current_asset_type == AssetType.PACKAGE:
            package_name = _row[8]
            package_version = _row[9]
            package_release = _row[10]
            package_size = _row[11]
            package_install_date = _row[12]
            package_type = _row[13]

            if package_size and isinstance(package_size, int):
                if package_type == PackageTypeEnum.WINDOWS.name:
                    package_size_str = sizeof_fmt(package_size * 1024)
                else:
                    package_size_str = sizeof_fmt(package_size)

            else:
                package_size_str = "N/A"

            if type(package_install_date) != str:
                package_install_date = package_install_date.strftime(ASSET_DATETIMEFORMAT)

            o = {
                'package_name': package_name,
                'package_version': package_version if package_version else "N/A",
                'package_release': package_release if package_release else "N/A",
                'package_size': package_size_str,
                'package_install_date': str(package_install_date),
                'package_type': package_type
            }

            self.rows_dict.append(o)

        elif self.current_asset_type == AssetType.PRETTYWINHW or self.current_asset_type == AssetType.LSHW:
            pass

        elif self.current_asset_type == AssetType.DMI:
            handle = str(_row[8])
            dmi_type = str(_row[9])
            header = str(_row[10])
            key = str(_row[11])
            value = str(_row[12])

            o = {
                'handle': handle,
                'dmi_type': dmi_type,
                'header': header,
                'key': key,
                'value': value
            }

            self.rows_dict.append(o)

        elif self.current_asset_type == AssetType.PCI:
            domain = _row[8]
            bus = _row[9]
            slot = _row[10]
            func = _row[11]
            position = _row[12]
            subclass = _row[13]
            vendor = _row[14]
            _device = _row[15]
            revision = _row[16]

            o = {
                'domain': domain,
                'bus': bus,
                'slot': slot,
                'func': func,
                'position': position,
                'subclass': subclass,
                'vendor': vendor,
                'device': _device,
                'revision': revision
            }

            self.rows_dict.append(o)


class GenericReport(object):
    def __init__(self, name):
        self.number_of_pages = 0
        self.pdf_buffers = []
        self.section_number = 0
        self.name = name
        self.root_report = None
        self.children = []

    def add_child(self, report):
        report.root_report = self
        free_section_number = 0
        for child in self.children:
            if child.section_number > free_section_number:
                free_section_number = child.section_number

        free_section_number += 1

        report.section_number = free_section_number

        self.children.append(report)

    def get_section_number(self):
        if self.root_report:
            return self.root_report.get_section_number() + ".{}".format(self.section_number)
        else:
            return "{}".format(self.section_number)

    def add_buffer_to_report(self, _buffer):
        self.pdf_buffers.append(_buffer)

    def increase_page_count(self, canvas, doc):
        id(canvas)
        id(doc)
        self.number_of_pages += 1


class DeviceReport(GenericReport):
    def __init__(self, _device, report_settings, name):
        super(DeviceReport, self).__init__(name)

        self.device = _device

        # default settings
        self.report_settings = {
            "packages_selected": True,
            "licenses_selected": True,
            "installed_updates_selected": True,
            "avail_updates_selected": True,
            "hardware_report_selected": True,
            "lstopo_report_selected": True,
            "process_report_selected": True,
            "dmi_report_selected": True,
            "pci_report_selected": True
        }

        # update default settings with new settings
        for setting in report_settings:
            self.report_settings[setting] = report_settings[setting]

    def module_selected(self, assetrun):
        if AssetType(assetrun.run_type) == AssetType.PACKAGE:
            return self.report_settings["packages_selected"]
        elif AssetType(assetrun.run_type) == AssetType.HARDWARE:
            return self.report_settings["lstopo_report_selected"]
        elif AssetType(assetrun.run_type) == AssetType.LICENSE:
            return self.report_settings["licenses_selected"]
        elif AssetType(assetrun.run_type) == AssetType.UPDATE:
            return self.report_settings["installed_updates_selected"]
        elif AssetType(assetrun.run_type) == AssetType.PROCESS:
            return self.report_settings["process_report_selected"]
        elif AssetType(assetrun.run_type) == AssetType.PENDING_UPDATE:
            return self.report_settings["avail_updates_selected"]
        elif AssetType(assetrun.run_type) == AssetType.DMI:
            return self.report_settings["dmi_report_selected"]
        elif AssetType(assetrun.run_type) == AssetType.PCI:
            return self.report_settings["pci_report_selected"]
        elif AssetType(assetrun.run_type) == AssetType.PRETTYWINHW:
            return self.report_settings["hardware_report_selected"]
        elif AssetType(assetrun.run_type) == AssetType.LSHW:
            return self.report_settings["hardware_report_selected"]


class ReportGenerator(object):
    def __init__(self, _settings, _devices):
        self.data = ""

        # general settings stored under special key -1
        self.general_settings = _settings[-1]
        self.device_settings = _settings

        self.devices = _devices

        report_history_obj = ReportHistory()
        report_history_obj.save()

        self.report_id = report_history_obj.idx

        self.creation_date = datetime.datetime.now()

    def get_report_data(self):
        return self.data

    def get_report_type(self):
        pass

    def set_progress(self, progress):
        report_history_obj = ReportHistory.objects.get(idx=self.report_id)

        if report_history_obj.progress != progress:
            report_history_obj.progress = progress
            report_history_obj.save()

    @staticmethod
    def _get_data_for_user_group_overview():
        users = user.objects.all()
        data = []

        for _user in users:
            o = {
                'login': _user.login,
                'uid': _user.uid,
                'firstname': _user.first_name,
                'lastname': _user.last_name,
                'email': _user.email,
                'group': _user.group.groupname
            }

            secondary_groups_str = _user.group.groupname

            for _group in _user.secondary_groups.all():
                secondary_groups_str += ", {}".format(_group.groupname)

            roles_str = "N/A"

            for role in _user.roles.all():
                if roles_str == "N/A":
                    roles_str = role.name
                else:
                    roles_str = "{}, {}".format(roles_str, role.name)

            allowed_device_group_str = "All/Any"
            for allowed_device_group in _user.allowed_device_groups.all():
                new_allowed_device_group_str = "{}".format(allowed_device_group.name)

                if allowed_device_group_str == "All/Any":
                    allowed_device_group_str = new_allowed_device_group_str
                else:
                    allowed_device_group_str = "{}, {}".format(allowed_device_group_str, new_allowed_device_group_str)

            o['secondary_groups'] = secondary_groups_str
            o['roles'] = roles_str
            o['allowed_device_groups'] = allowed_device_group_str

            data.append(o)

        return data

    @staticmethod
    def _get_data_for_user_roles_overview():
        from initat.cluster.backbone.models.user import Role

        ac_to_str_dict = {
            AC_READONLY: "Read-only",
            AC_MODIFY: "Modify",
            AC_CREATE: "Modify, Create",
            AC_FULL: "Modify, Create, Delete"
        }

        data = []

        for role in Role.objects.all():
            o = {
                'name': role.name,
                'description': role.description if role.description else "N/A"
            }

            permission_str = "N/A"

            index = 1
            for role_permission in role.rolepermission_set.all():
                permission_name = role_permission.csw_permission.name

                new_permission_str = "{}. {} [{}]".format(index, permission_name, ac_to_str_dict[role_permission.level])

                if permission_str == "N/A":
                    permission_str = new_permission_str
                else:
                    permission_str += "\n{}".format(new_permission_str)

                index += 1

            for role_object_permission in role.roleobjectpermission_set.all():
                permission_name = role_object_permission.csw_object_permission.csw_permission.name
                object_pk = role_object_permission.csw_object_permission.object_pk

                content_type = role_object_permission.csw_object_permission.csw_permission.content_type
                _object = content_type.get_object_for_this_type(pk=object_pk)

                new_permission_str = "{}: {} for {} [{}]".format(index, permission_name, str(_object),
                                                                 ac_to_str_dict[role_object_permission.level])

                if permission_str == "N/A":
                    permission_str = new_permission_str
                else:
                    permission_str += "\n{}".format(new_permission_str)

                index += 1

            o['permission'] = permission_str

            data.append(o)

        return data

    @staticmethod
    def _get_data_for_network_report():
        data = []
        for _network in network.objects.all():
            o = {
                'id': _network.identifier,
                'network': _network.network,
                'netmask': _network.netmask,
                'broadcast': _network.broadcast,
                'gateway': _network.gateway,
                'gwprio': _network.gw_pri,
                'num_ips': _network.num_ip(),
                'network_type': _network.network_type.description
            }

            data.append(o)

        return data

    @staticmethod
    def _get_data_for_sub_network_report(_network):
        data = []
        for net_ip in _network.net_ip_set.all():
            o = {
                'ip': net_ip.ip,
                'netdevname': net_ip.netdevice.devname,
                'devname': net_ip.netdevice.device.full_name
            }

            data.append(o)

        return data


class ColoredRule(Rule):
    def __init__(self, pos, width, thickness=1.0, report=None, hexcolor=None):
        super(ColoredRule, self).__init__(pos, width, thickness, report)

        self.hexcolor = hexcolor

    def generate(self, row):
        return ColoredRule(self.pos, self.width, self.height, self.report, self.hexcolor)

    def render(self, offset, canvas):
        leftmargin = self.report.leftmargin
        canvas.saveState()
        canvas.setLineWidth(self.height)
        canvas.setFillColor(self.hexcolor)
        canvas.setStrokeColor(self.hexcolor)
        canvas.line(self.pos[0] + leftmargin,
                    -1 * (self.pos[1] + offset + self.height / 2),
                    self.pos[0] + self.width + leftmargin,
                    -1 * (self.pos[1] + offset + self.height / 2))

        canvas.restoreState()


class PDFReportGenerator(ReportGenerator):
    def __init__(self, _settings, _devices):
        super(PDFReportGenerator, self).__init__(_settings, _devices)

        # logo and styling options/settings
        system_device = None
        for _device in device.objects.all():
            if _device.is_cluster_device_group():
                system_device = _device
                break

        report_logo = system_device.device_variable_set.filter(name="__REPORT_LOGO__")
        self.logo_width = None
        self.logo_height = None
        self.logo_buffer = BytesIO()
        if report_logo:
            report_logo = report_logo[0]
            data = base64.b64decode(report_logo.val_blob)
            self.logo_buffer.write(data)
            im = PILImage.open(self.logo_buffer)
            self.logo_width = im.size[0]
            self.logo_height = im.size[1]

            drawheight_max = 45
            drawwidth_max = 105

            ratio = float(self.logo_width) / float(self.logo_height)

            while self.logo_width > drawwidth_max or self.logo_height > drawheight_max:
                self.logo_width -= ratio * 1
                self.logo_height -= 1

        else:
            # generate a simple placeholder image
            self.logo_height = 42
            self.logo_width = 103
            logo = PILImage.new('RGB', (255, 255), "white")  # create a new white image
            logo.save(self.logo_buffer, format="BMP")

        self.page_format = eval(self.general_settings["pdf_page_format"])
        self.margin = 36

        self.standard_font = "SourceSansPro-Regular"
        self.bold_font = "SourceSansPro-Bold"

        self.cluster_id = "Unknown"
        cluster_id_var = system_device.device_variable_set.filter(name="CLUSTER_ID")
        if cluster_id_var:
            cluster_id_var = cluster_id_var[0]
            self.cluster_id = cluster_id_var.val_str

        self.cluster_name = "Unknown"
        cluster_name_var = system_device.device_variable_set.filter(name="CLUSTER_NAME")
        if cluster_name_var:
            cluster_name_var = cluster_name_var[0]
            self.cluster_name = cluster_name_var.val_str

        # Storage dicts / lists
        self.current_page_num = 0

        self.reports = []

    def __scale_logo(self, drawheight_max=None, drawwidth_max=None):
        im = PILImage.open(self.logo_buffer)
        logo_width = im.size[0]
        logo_height = im.size[1]

        ratio = float(logo_width) / float(logo_height)

        while (drawwidth_max and logo_width > drawwidth_max) or (drawheight_max and logo_height > drawheight_max):
            logo_width -= ratio * 1
            logo_height -= 1

        return logo_width, logo_height

    def __config_report_helper(self, header, header_names_left, header_names_right, rpt, data):
        available_width = self.page_format[0] - (self.margin * 2)

        header_list = [Element((0, 0), (self.bold_font, 16), text=header)]

        detail_list = []

        position = 0
        rule_position = 42

        header_font_size = 12
        normal_font_size = 6

        if self.general_settings['pdf_page_format'] == "A4" or self.general_settings['pdf_page_format'] == "letter":
            header_font_size = 8
            rule_position = 36

        for header_name, key, avail_width_percentage in header_names_left:
            s_new = ""
            wrap_idx = 0

            for i in range(len(header_name)):
                width = stringWidth(header_name[wrap_idx:i + 1], self.standard_font, header_font_size)
                if ((width / available_width) * 101.5) > avail_width_percentage:
                    wrap_idx = i
                    s_new += "\n"
                s_new += header_name[i]

            header_list.append(Element((position, 24), (self.standard_font, header_font_size), text=s_new))
            detail_list.append(Element((position, 0), (self.standard_font, normal_font_size), key=key))

            position += available_width * (avail_width_percentage / 100.0)

            for _dict in data:
                s = str(_dict[key])

                s_new_comps = ""

                comps = s.split("\n")
                for comp in comps:
                    s_new = ""
                    wrap_idx = 0

                    for i in range(len(comp)):
                        width = stringWidth(s[wrap_idx:i+1], self.standard_font, normal_font_size)
                        if ((width / available_width) * 101.5) > avail_width_percentage:
                            wrap_idx = i
                            s_new += "\n"
                        s_new += comp[i]

                    if s_new_comps:
                        s_new_comps += "\n{}".format(s_new)
                    else:
                        s_new_comps = s_new

                _dict[key] = s_new_comps

        position = self.page_format[0] - (self.margin * 2)

        for header_name, key, avail_width_percentage in reversed(header_names_right):
            s_new = ""
            wrap_idx = 0

            for i in range(len(header_name)):
                width = stringWidth(header_name[wrap_idx:i + 1], self.standard_font, header_font_size)
                if ((width / available_width) * 101.5) > avail_width_percentage:
                    wrap_idx = i
                    s_new += "\n"
                s_new += header_name[i]

            header_list.append(
                Element(
                    (position, 24),
                    (self.standard_font, header_font_size),
                    text=s_new,
                    align="right"
                )
            )
            detail_list.append(
                Element(
                    (position, 0),
                    (self.standard_font, normal_font_size),
                    key=key,
                    align="right"
                )
            )

            position -= available_width * (avail_width_percentage / 100.0)

            for _dict in data:
                s = str(_dict[key])

                s_new_comps = ""

                comps = s.split("\n")
                for comp in comps:
                    s_new = ""
                    wrap_idx = 0

                    for i in range(len(comp)):
                        width = stringWidth(s[wrap_idx:i+1], self.standard_font, normal_font_size)
                        if ((width / available_width) * 101.5) > avail_width_percentage:
                            wrap_idx = i
                            s_new += "\n"
                        s_new += comp[i]

                    if s_new_comps:
                        s_new_comps += "\n{}".format(s_new)
                    else:
                        s_new_comps = s_new

                _dict[key] = s_new_comps

        header_list.append(ColoredRule((0, rule_position), self.page_format[0] - (self.margin * 2), thickness=1.5,
            hexcolor=HexColor(0xBDBDBD)))
        detail_list.append(ColoredRule((0, 0), self.page_format[0] - (self.margin * 2), thickness=0.1,
            hexcolor=HexColor(0xBDBDBD)))

        rpt.pageheader = Band(header_list)
        rpt.detailband = Band(detail_list)

    def __generate_general_device_overview_report(self, root_report):
        _buffer = BytesIO()
        canvas = Canvas(_buffer, self.page_format)

        data = _generate_hardware_info_data_dict(self.devices, self.general_settings["assetbatch_selection_mode"])

        data = sorted(data, key=lambda k: k['group'])
        if data:
            report = GenericReport("Device Overview")
            root_report.add_child(report)

            section_number = report.get_section_number()

            rpt = PollyReportsReport(data)
            rpt.groupheaders = [Band([Element((0, 4), (self.bold_font, 10),
                                              getvalue=lambda x: x['group'],
                                              format=lambda x: "Group: {}".format(x)), ],
                                     getvalue=lambda x: x["group"])]

            header_names_left = [("Device Name", "name", 10.0),
                                 ("CPU Info", "cpu", 18.0),
                                 ("GPU Info", "gpu", 18.0),
                                 ("Memory Info", "memory", 18.0),
                                 ("HDD Info", "hdd", 18.0),
                                 ("Partition Info", "partition", 18.0)
                                 ]
            header_names_right = []

            self.__config_report_helper("{} Device Overview".format(section_number), header_names_left,
                                        header_names_right, rpt, data)

            rpt.generate(canvas)
            canvas.save()
            report.number_of_pages += rpt.pagenumber
            report.add_buffer_to_report(_buffer)
            self.current_page_num += rpt.pagenumber

    def __generate_network_report(self, root_report):
        _buffer = BytesIO()
        canvas = Canvas(_buffer, self.page_format)

        data = self._get_data_for_network_report()

        data = sorted(data, key=lambda k: k['id'])
        if data:
            report = GenericReport("Networks")
            root_report.add_child(report)

            section_number = report.get_section_number()

            rpt = PollyReportsReport(data)
            rpt.groupheaders = [Band([Element((0, 4), (self.bold_font, 10),
                                              getvalue=lambda x: x['id'][0],
                                              format=lambda x: "Networks starting with: {}".format(x)), ],
                                     getvalue=lambda x: x["id"][0])]

            header_names_left = [("Identifier", "id", 12.5),
                                 ("Network", "network", 12.5),
                                 ("Netmask", "netmask", 12.5),
                                 ("Broadcast", "broadcast", 12.5),
                                 ("Gateway", "gateway", 12.5),
                                 ('GW Priority', "gwprio", 12.5),
                                 ("#IPs", "num_ips", 12.5),
                                 ("Network Type", "network_type", 12.5)]

            header_names_right = []

            self.__config_report_helper("{} Network Overview".format(section_number),
                                        header_names_left,
                                        header_names_right, rpt, data)

            rpt.generate(canvas)
            canvas.save()
            report.number_of_pages += rpt.pagenumber
            report.add_buffer_to_report(_buffer)
            self.current_page_num += rpt.pagenumber

            for _network in network.objects.all():
                self.__generate_sub_network_report(_network, report)

    def __generate_sub_network_report(self, _network, root_report):
        _buffer = BytesIO()
        canvas = Canvas(_buffer, self.page_format)

        data = self._get_data_for_sub_network_report(_network)

        data = sorted(data, key=lambda k: k['ip'])
        if data:
            report = GenericReport(_network.identifier)
            root_report.add_child(report)
            section_number = report.get_section_number()

            rpt = PollyReportsReport(data)

            header_names_left = [("IP", "ip", 33.0),
                                 ("Net-Device", "netdevname", 33.0),
                                 ("Device", "devname", 34.0)]

            header_names_right = []

            self.__config_report_helper("{} Network: {}".format(section_number, _network.identifier),
                                        header_names_left,
                                        header_names_right, rpt, data)

            rpt.generate(canvas)
            canvas.save()
            report.number_of_pages += rpt.pagenumber
            report.add_buffer_to_report(_buffer)
            self.current_page_num += rpt.pagenumber

    def __generate_device_report(self, _device, report_settings, root_report):
        device_report = DeviceReport(_device, report_settings, _device.full_name)
        root_report.add_child(device_report)

        # create generic overview page
        self.__generate_device_overview_report(_device, report_settings, device_report)
        self.__generate_device_assetrun_reports(_device, report_settings, device_report)

    def __generate_device_overview_report(self, _device, report_settings, root_report):
        report = DeviceReport(_device, report_settings, "Overview")
        root_report.add_child(report)

        section_number = report.get_section_number()

        _buffer = BytesIO()
        doc = SimpleDocTemplate(_buffer,
                                pagesize=self.page_format,
                                rightMargin=0,
                                leftMargin=0,
                                topMargin=10,
                                bottomMargin=25)
        elements = []

        style_sheet = getSampleStyleSheet()

        available_width = self.page_format[0] - 60

        paragraph_header = Paragraph('<font face="{}" size="16">{} Overview for {}</font>'.format(
            self.bold_font, section_number, _device.name), style_sheet["BodyText"])

        data = [[paragraph_header]]

        t_head = Table(data,
                       colWidths=(available_width),
                       rowHeights=[35],
                       style=[('LEFTPADDING', (0, 0), (-1, -1), 0),
                              ('RIGHTPADDING', (0, 0), (-1, -1), 0),])

        body_data = []

        # FQDN
        data = [[_device.full_name]]

        text_block = Paragraph('<b>FQDN:</b>', style_sheet["BodyText"])
        t = Table(data, colWidths=(available_width * 0.83),
                  style=[]
                  )
        body_data.append((text_block, t))

        # Device Groups
        data = [[_device.device_group_name()]]

        text_block = Paragraph('<b>DeviceGroup:</b>', style_sheet["BodyText"])
        t = Table(data, colWidths=(available_width * 0.83),
                  style=[]
                  )

        body_data.append((text_block, t))

        # Device Class
        data = [[_device.device_class.name]]

        text_block = Paragraph('<b>DeviceClass:</b>', style_sheet["BodyText"])
        t = Table(data, colWidths=(available_width * 0.83),
                  style=[]
                  )

        body_data.append((text_block, t))

        # ComCapabilites
        str_to_use = "N/A"

        for com_cap in _device.com_capability_list.all():
            if str_to_use == "N/A":
                str_to_use = com_cap.name
            else:
                str_to_use += ", {}".format(com_cap.name)
        data = [[str_to_use]]

        text_block = Paragraph('<b>ComCapabilities:</b>', style_sheet["BodyText"])
        t = Table(data, colWidths=(available_width * 0.83),
                  style=[]
                  )

        body_data.append((text_block, t))

        # Ip info
        str_to_use = "N/A"
        for _ip in _device.all_ips():
            if str_to_use == "N/A":
                str_to_use = str(_ip)
            else:
                str_to_use += ", {}".format(str(_ip))
        data = [[str_to_use]]

        text_block = Paragraph('<b>IP Info:</b>', style_sheet["BodyText"])
        t = Table(data, colWidths=(available_width * 0.83),
                  style=[]
                  )

        body_data.append((text_block, t))

        # SNMP Schemes
        str_to_use = "N/A"
        for _snmp_scheme in _device.snmp_schemes.all():
            if str_to_use == "N/A":
                str_to_use = str(_snmp_scheme)
            else:
                str_to_use += ", {}".format(str(_snmp_scheme))
        data = [[str_to_use]]

        text_block = Paragraph('<b>SNMP Scheme:</b>', style_sheet["BodyText"])
        t = Table(data, colWidths=(available_width * 0.83),
                  style=[]
                  )

        body_data.append((text_block, t))

        # SNMP Info
        str_to_use = "N/A"
        data = [[str_to_use]]

        text_block = Paragraph('<b>SNMP Info:</b>', style_sheet["BodyText"])
        t = Table(data, colWidths=(available_width * 0.83),
                  style=[]
                  )

        body_data.append((text_block, t))

        # Device Categories
        str_to_use = "N/A"
        for _category in _device.categories.all():
            if str_to_use == "N/A":
                str_to_use = _category.name
            else:
                str_to_use += ", {}".format(_category.name)
        data = [[str_to_use]]

        text_block = Paragraph('<b>Categories:</b>', style_sheet["BodyText"])
        t = Table(data, colWidths=(available_width * 0.83),
                  style=[]
                  )

        body_data.append((text_block, t))

        t_body = Table(body_data, colWidths=(available_width * 0.15, available_width * 0.85),
                       style=[('VALIGN', (0, 0), (0, -1), 'MIDDLE'),
                              ('GRID', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                              ('BOX', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                              ])

        elements.append(t_head)
        elements.append(Spacer(1, 30))
        elements.append(t_body)

        doc.build(elements, onFirstPage=report.increase_page_count, onLaterPages=report.increase_page_count)
        report.add_buffer_to_report(_buffer)

    def __generate_device_assetrun_reports(self, _device, report_settings, root_report):
        assetruns = _select_assetruns_for_device(_device, self.general_settings["assetbatch_selection_mode"])

        row_collector = RowCollector()

        # Hardware report is generated last
        hardware_report_ar = None

        sorted_runs = {}

        for ar in assetruns:
            if AssetType(ar.run_type) == AssetType.PACKAGE:
                sorted_runs[0] = ar
            elif AssetType(ar.run_type) == AssetType.LICENSE:
                sorted_runs[1] = ar
            elif AssetType(ar.run_type) == AssetType.UPDATE:
                sorted_runs[2] = ar
            elif AssetType(ar.run_type) == AssetType.PENDING_UPDATE:
                sorted_runs[3] = ar
            elif AssetType(ar.run_type) == AssetType.PROCESS:
                sorted_runs[4] = ar
            elif AssetType(ar.run_type) == AssetType.PRETTYWINHW or AssetType(ar.run_type) == AssetType.LSHW:
                sorted_runs[5] = ar
            elif AssetType(ar.run_type) == AssetType.DMI:
                sorted_runs[6] = ar
            elif AssetType(ar.run_type) == AssetType.PCI:
                sorted_runs[7] = ar
            elif AssetType(ar.run_type) == AssetType.HARDWARE:
                sorted_runs[8] = ar

        for idx in sorted_runs:
            _buffer = BytesIO()

            canvas = Canvas(_buffer, self.page_format)

            ar = sorted_runs[idx]

            row_collector.reset()
            row_collector.current_asset_type = AssetType(ar.run_type)
            generate_csv_entry_for_assetrun(ar, row_collector.collect)

            if AssetType(ar.run_type) == AssetType.UPDATE:
                data = row_collector.rows_dict[1:]
                data = sorted(data, key=lambda k: k['update_status'])

                heading = "System Updates"

                report = DeviceReport(_device, report_settings, heading)
                if not report.report_settings['installed_updates_selected']:
                    continue

                root_report.add_child(report)

                section_number = report.get_section_number()

                rpt = PollyReportsReport(data)

                if data:
                    rpt.groupheaders = [Band([Element((0, 4), (self.bold_font, 10),
                                                      getvalue=lambda x: x['update_status'],
                                                      format=lambda x: "Updates with status: {}".format(x)), ],
                                             getvalue=lambda x: x["update_status"]), ]
                else:
                    mock_object = {
                        "update_name": "",
                        "install_date": "",
                        "update_status": "",
                    }
                    data.append(mock_object)

                header_names_left = [("Update Name", "update_name", 70.0),
                                     ("Install Status", "update_status", 15.0)]
                header_names_right = [("Install Date", "install_date", 15.0)]

                self.__config_report_helper("{} {} for {}".format(section_number, heading, _device.full_name),
                                            header_names_left, header_names_right, rpt, data)

                rpt.generate(canvas)
                report.number_of_pages += rpt.pagenumber

            elif AssetType(ar.run_type) == AssetType.LICENSE:
                data = row_collector.rows_dict[1:]

                heading = "Active Licenses"

                report = DeviceReport(_device, report_settings, heading)
                if not report.report_settings['licenses_selected']:
                    continue

                root_report.add_child(report)

                section_number = report.get_section_number()

                rpt = PollyReportsReport(data)

                if not data:
                    mock_object = {
                        "license_name": "",
                        "license_key": ""
                    }
                    data.append(mock_object)

                header_names_left = [("License Name", "license_name", 50.0),
                                     ("License Key", "license_key", 50.0)]
                header_names_right = []

                self.__config_report_helper("{} {} for {}".format(section_number, heading, _device.full_name),
                                            header_names_left, header_names_right, rpt, data)

                rpt.generate(canvas)
                report.number_of_pages += rpt.pagenumber

            elif AssetType(ar.run_type) == AssetType.PACKAGE:
                from initat.cluster.backbone.models import asset

                try:
                    packages = asset.get_packages_for_ar(ar)
                except Exception as e:
                    logger.info("PDF generation for packages failed, error was: {}".format(str(e)))
                    packages = []

                heading = "Installed Software"

                report = DeviceReport(_device, report_settings, heading)
                if not report.report_settings['packages_selected']:
                    continue

                root_report.add_child(report)

                section_number = report.get_section_number()

                data = [package.get_as_row() for package in packages]
                data = sorted(data, key=lambda k: k['package_name'])

                rpt = PollyReportsReport(data)

                if data:
                    rpt.groupheaders = [Band([Element((0, 4), (self.bold_font, 10),
                                                      getvalue=lambda x: x['package_name'][0],
                                                      format=lambda x: "Packages starting with: {}".format(x)), ],
                                             getvalue=lambda x: x["package_name"][0]), ]
                else:
                    mock_object = {
                        "package_name": "",
                        "package_version": "",
                        "package_release": "",
                        "package_size": "",
                        "package_install_date": ""
                    }
                    data.append(mock_object)

                header_names_left = [("Name", "package_name", 40.0),
                                     ("Version", "package_version", 15.00),
                                     ("Release", "package_release", 15.00)]
                header_names_right = [("Size", "package_size", 15.00),
                                      ("Install Date", "package_install_date", 15.00)]

                self.__config_report_helper("{} {} for {}".format(section_number, heading, _device.full_name),
                                            header_names_left, header_names_right, rpt, data)

                rpt.generate(canvas)
                report.number_of_pages += rpt.pagenumber

            elif AssetType(ar.run_type) == AssetType.PENDING_UPDATE:
                if not report.report_settings['avail_updates_selected']:
                    continue

                data = row_collector.rows_dict[1:]
                data = sorted(data, key=lambda k: k['update_name'])

                heading = "Updates ready for install"

                report = DeviceReport(_device, report_settings, heading)
                root_report.add_child(report)

                section_number = report.get_section_number()

                rpt = PollyReportsReport(data)

                if data:
                    rpt.groupheaders = [Band([Element((0, 4), (self.bold_font, 10),
                                                      getvalue=lambda x: x['update_name'][0].upper(),
                                                      format=lambda x: "Updates starting with: {}".format(x)), ],
                                             getvalue=lambda x: x["update_name"][0].upper()), ]
                else:
                    mock_object = {
                        "update_name": "",
                        "update_version": "",
                        "update_optional": ""
                    }
                    data.append(mock_object)

                header_names_left = [("Update Name", "update_name", 33.33),
                                     ("Version", "update_version", 33.33),
                                     ("Optional", "update_optional", 33.33)]
                header_names_right = []

                self.__config_report_helper("{} {} for {}".format(section_number, heading, _device.full_name),
                                            header_names_left, header_names_right, rpt, data)

                rpt.generate(canvas)
                report.number_of_pages += rpt.pagenumber

            elif AssetType(ar.run_type) == AssetType.HARDWARE:
                if not report.report_settings['lstopo_report_selected']:
                    continue

                data = row_collector.rows_dict[1:]
                data = sorted(data, key=lambda k: k['hardware_depth'])

                heading = "Lstopo Information"

                report = DeviceReport(_device, report_settings, heading)
                root_report.add_child(report)

                section_number = report.get_section_number()

                rpt = PollyReportsReport(data)

                if not data:
                    mock_object = {
                        "hardware_node_type": "",
                        "hardware_depth": "",
                        "hardware_attributes": "",
                    }
                    data.append(mock_object)

                header_names_left = [("Node Type", "hardware_node_type", 15.00),
                                     ("Depth", "hardware_depth", 15.00),
                                     ("Attributes", "hardware_attributes", 70.00)]
                header_names_right = []

                self.__config_report_helper("{} {} for {}".format(section_number, heading, _device.full_name),
                                            header_names_left, header_names_right, rpt, data)

                rpt.generate(canvas)
                report.number_of_pages += rpt.pagenumber

            elif AssetType(ar.run_type) == AssetType.PROCESS:
                if not report.report_settings['process_report_selected']:
                    continue

                data = row_collector.rows_dict[1:]
                data = sorted(data, key=lambda k: k['process_name'])

                heading = "Process Information"

                report = DeviceReport(_device, report_settings, heading)
                root_report.add_child(report)

                section_number = report.get_section_number()

                rpt = PollyReportsReport(data)

                if data:
                    rpt.groupheaders = [Band([Element((0, 4), (self.bold_font, 10),
                                                      getvalue=lambda x: x['process_name'][0],
                                                      format=lambda x: "Processes starting with: {}".format(x)), ],
                                             getvalue=lambda x: x["process_name"][0]), ]
                else:
                    mock_object = {
                        "process_name": "",
                        "process_id": "",
                    }
                    data.append(mock_object)

                header_names_left = [("Process Name", "process_name", 50.0),
                                     ("PID", "process_id", 50.0)]
                header_names_right = []

                self.__config_report_helper("{} {} for {}".format(section_number, heading, _device.full_name),
                                            header_names_left, header_names_right, rpt, data)

                rpt.generate(canvas)
                report.number_of_pages += rpt.pagenumber

            elif AssetType(ar.run_type) == AssetType.DMI:
                if not report.report_settings['dmi_report_selected']:
                    continue

                heading = "Hardware Details"

                data = row_collector.rows_dict[1:]

                report = DeviceReport(_device, report_settings, heading)
                root_report.add_child(report)

                section_number = report.get_section_number()

                rpt = PollyReportsReport(data)

                if not data:
                    mock_object = {
                        "handle": "",
                        "dmi_type": "",
                        "header": "",
                        "key": "",
                        "value": ""
                    }
                    data.append(mock_object)

                header_names_left = [("Handle", "handle", 8.0),
                                     ("Type", "dmi_type", 8.0),
                                     ("Header", "header", 15.0),
                                     ("Key", "key", 15.0),
                                     ("Value", "value", 54.0)]
                header_names_right = []

                self.__config_report_helper("{} {} for {}".format(section_number, heading, _device.full_name),
                                            header_names_left, header_names_right, rpt, data)

                rpt.generate(canvas)
                report.number_of_pages += rpt.pagenumber

            elif AssetType(ar.run_type) == AssetType.PCI:
                if not report.report_settings['pci_report_selected']:
                    continue

                data = row_collector.rows_dict[1:]

                heading = "PCI Details"

                report = DeviceReport(_device, report_settings, heading)
                root_report.add_child(report)

                section_number = report.get_section_number()

                rpt = PollyReportsReport(data)

                if not data:
                    mock_object = {
                        'domain': "",
                        'bus': "",
                        'slot': "",
                        'func': "",
                        'position': "",
                        'subclass': "",
                        'vendor': "",
                        'device': "",
                        'revision': ""
                    }

                    data.append(mock_object)

                header_names_left = [("Domain", "domain", 7.5),
                                     ("Bus", "bus", 5.0),
                                     ("Slot", "slot", 5.0),
                                     ("Func", "func", 5.0),
                                     ("Position", "position", 10.0),
                                     ("Subclass", "subclass", 10.0),
                                     ("Vendor", "vendor", 28.75),
                                     ("Device", "device", 28.75)]
                header_names_right = []

                self.__config_report_helper("{} {} for {}".format(section_number, heading, _device.full_name),
                                            header_names_left, header_names_right, rpt, data)

                rpt.generate(canvas)
                report.number_of_pages += rpt.pagenumber

            elif ar.run_type == AssetType.PRETTYWINHW or ar.run_type == AssetType.LSHW:
                if not report.report_settings['hardware_report_selected']:
                    continue

                hardware_report_ar = ar
                continue

            canvas.save()
            report.add_buffer_to_report(_buffer)

        if hardware_report_ar:
            self.__generate_hardware_report(hardware_report_ar, _device, report_settings, root_report)

    def __generate_hardware_report(self, hardware_report_ar, _device, report_settings, root_report):
        report = DeviceReport(_device, report_settings, "Hardware Report")
        root_report.add_child(report)

        section_number = report.get_section_number()

        _buffer = BytesIO()
        doc = SimpleDocTemplate(_buffer,
                                pagesize=self.page_format,
                                rightMargin=0,
                                leftMargin=0,
                                topMargin=10,
                                bottomMargin=25)
        elements = []

        style_sheet = getSampleStyleSheet()
        style_sheet.add(ParagraphStyle(name='courier',
                                       fontName='Courier',
                                       fontSize=12))

        available_width = self.page_format[0] - 60

        data = [["Name", "Cores"]]
        for cpu in hardware_report_ar.asset_batch.cpus.all():
            data.append([Paragraph(str(cpu.cpuname), style_sheet["BodyText"]),
                         Paragraph(str(cpu.numberofcores), style_sheet["BodyText"])])

        p0_1 = Paragraph('<b>CPUs:</b>', style_sheet["BodyText"])

        t_1 = Table(data,
                    colWidths=(available_width * 0.88 * 0.50,
                               available_width * 0.88 * 0.50),
                    style=[('GRID', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                           ('BOX', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD))])

        data = [["Name", "Driver Version"]]
        for gpu in hardware_report_ar.asset_batch.gpus.all():
            data.append([Paragraph(str(gpu.gpuname), style_sheet["BodyText"]),
                         Paragraph(str(gpu.driverversion), style_sheet["BodyText"])])

        p0_2 = Paragraph('<b>GPUs:</b>', style_sheet["BodyText"])
        t_2 = Table(data,
                    colWidths=(available_width * 0.88 * 0.50,
                               available_width * 0.88 * 0.50),
                    style=[('GRID', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                           ('BOX', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                           ])

        data = [["Name", "Serialnumber", "Size"]]
        for hdd in hardware_report_ar.asset_batch.hdds.all():
            data.append([Paragraph(str(hdd.name), style_sheet["BodyText"]),
                         Paragraph(str(hdd.serialnumber), style_sheet["courier"]),
                         Paragraph(sizeof_fmt(hdd.size), style_sheet["BodyText"])])

        p0_3 = Paragraph('<b>HDDs:</b>', style_sheet["BodyText"])
        t_3 = Table(data,
                    colWidths=(available_width * 0.88 * 0.33,
                               available_width * 0.88 * 0.34,
                               available_width * 0.88 * 0.33),
                    style=[('GRID', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                           ('BOX', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                           ])

        data = [["Name", "Size", "Free", "Graph"]]
        for partition in hardware_report_ar.asset_batch.partitions.all():
            d = Drawing(10, 10)
            r = Rect(0, 0, 130, 12)
            r.fillColor = colors.red
            d.add(r)

            if partition.size is not None and partition.free is not None:
                free_length = int((float(partition.free) / float(partition.size)) * 130)
                free_start = 130 - free_length

                r = Rect(free_start, 0, free_length, 12)
                r.fillColor = colors.green
                d.add(r)
            else:
                d = Paragraph("N/A", style_sheet["BodyText"])

            data.append([Paragraph(str(partition.name), style_sheet["BodyText"]),
                         Paragraph(sizeof_fmt(partition.size), style_sheet["BodyText"]),
                         Paragraph(sizeof_fmt(partition.free), style_sheet["BodyText"]),
                         d])

        p0_4 = Paragraph('<b>Partitions:</b>', style_sheet["BodyText"])
        t_4 = Table(data,
                    colWidths=(available_width * 0.88 * 0.25,
                               available_width * 0.88 * 0.25,
                               available_width * 0.88 * 0.25,
                               available_width * 0.88 * 0.25),
                    style=[('GRID', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                           ('BOX', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                           ])

        data = [["Banklabel", "Formfactor", "Memorytype", "Manufacturer", "Capacity"]]
        for memory_module in hardware_report_ar.asset_batch.memory_modules.all():

            data.append([Paragraph(str(memory_module.banklabel), style_sheet["BodyText"]),
                         Paragraph(str(memory_module.get_name_of_form_factor()), style_sheet["BodyText"]),
                         Paragraph(str(memory_module.get_name_of_memory_type()), style_sheet["BodyText"]),
                         Paragraph(str(memory_module.manufacturer), style_sheet["BodyText"]),
                         Paragraph(sizeof_fmt(memory_module.capacity), style_sheet["BodyText"])])

        p0_5 = Paragraph('<b>Memory Modules:</b>', style_sheet["BodyText"])
        t_5 = Table(data,
                    colWidths=(available_width * 0.88 * 0.2,
                               available_width * 0.88 * 0.2,
                               available_width * 0.88 * 0.2,
                               available_width * 0.88 * 0.2,
                               available_width * 0.88 * 0.2),
                    style=[('GRID', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                           ('BOX', (0, 0), (-1, -1), 0.35, HexColor(0xBDBDBD)),
                           ])

        data = [[p0_1, t_1],
                [p0_2, t_2],
                [p0_3, t_3],
                [p0_4, t_4],
                [p0_5, t_5]]

        t_body = Table(data, colWidths=(available_width * 0.10, available_width * 0.90),
                       style=[('VALIGN', (0, 0), (0, -1), 'MIDDLE')])

        p_h = Paragraph('<font face="{}" size="16">{} Hardware Report for {}</font>'.format(
            self.bold_font,
            section_number,
            hardware_report_ar.asset_batch.device.name),
            style_sheet["BodyText"])

        data = [[p_h]]

        t_head = Table(data,
                       colWidths=(available_width),
                       rowHeights=[35],
                       style=[])

        elements.append(t_head)
        elements.append(Spacer(1, 30))
        elements.append(t_body)

        doc.build(elements, onFirstPage=report.increase_page_count, onLaterPages=report.increase_page_count)

        report.add_buffer_to_report(_buffer)

    def __generate_front_page(self, number_of_pages):
        _buffer = BytesIO()
        doc = SimpleDocTemplate(_buffer,
                                pagesize=self.page_format,
                                rightMargin=0,
                                leftMargin=0,
                                topMargin=0,
                                bottomMargin=0)
        elements = []

        style_sheet = getSampleStyleSheet()
        style_sheet.add(ParagraphStyle(name='red font', textColor=colors.red))
        style_sheet.add(ParagraphStyle(name='green font', textColor=colors.green))
        style_sheet.add(ParagraphStyle(name="heading_1",
                                       fontName="SourceSansPro-Regular",
                                       fontSize=36,
                                       alignment=TA_CENTER))
        style_sheet.add(ParagraphStyle(name="heading_2",
                                       fontName="SourceSansPro-Regular",
                                       fontSize=14,
                                       alignment=TA_CENTER))
        style_sheet.add(ParagraphStyle(name="heading_3",
                                       fontName="SourceSansPro-Regular",
                                       fontSize=16,
                                       alignment=TA_CENTER))
        style_sheet.add(ParagraphStyle(name="heading_4",
                                       fontName="SourceSansPro-Regular",
                                       fontSize=11,
                                       alignment=TA_CENTER))

        available_width = self.page_format[0] - (self.margin * 4)

        data = [
            [Paragraph("Asset report", style_sheet['heading_1'])],
            [Paragraph("Report #{}".format(self.report_id), style_sheet['heading_2'])]
            ]

        h, w = self.__scale_logo(drawheight_max=23 * mm)

        logo = Image(self.logo_buffer)
        logo.drawHeight = w
        logo.drawWidth = h

        data.append([logo])

        data.append([Paragraph("{}".format(self.cluster_name), style_sheet['heading_3'])])
        data.append([Paragraph("This report consists of {} pages".format(number_of_pages), style_sheet['heading_4'])])

        data.append([Paragraph("General Content:", style_sheet['heading_2'])])

        general_content_str = "N/A"

        if self.general_settings['network_report_overview_module_selected']:
            var = "Network Overview"
            if general_content_str == "N/A":
                general_content_str = var
            else:
                general_content_str = "{}, {}".format(general_content_str, var)

        if self.general_settings['general_device_overview_module_selected']:
            var = "Device Overview"
            if general_content_str == "N/A":
                general_content_str = var
            else:
                general_content_str = "{}, {}".format(general_content_str, var)

        if self.general_settings['user_group_overview_module_selected']:
            var = "User/Group Overview"
            if general_content_str == "N/A":
                general_content_str = var
            else:
                general_content_str = "{}, {}".format(general_content_str, var)

        data.append([Paragraph(general_content_str, style_sheet['heading_4'])])

        selected_devices_str = "N/A"

        data.append([Paragraph("{} Devices:".format(len(self.devices)), style_sheet['heading_2'])])

        for _device in self.devices:
            var = _device.full_name
            if selected_devices_str == "N/A":
                selected_devices_str = var
            else:
                selected_devices_str = "{}, {}".format(selected_devices_str, var)

        data.append([Paragraph(selected_devices_str, style_sheet['heading_4'])])

        t_head = Table(data, colWidths=[available_width], rowHeights=[70, 45, 150, 15, 20, 50, 16, 50, None],
                       style=[
                           ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                           ('LEFTPADDING', (0, 0), (-1, -1), 0),
                           ('RIGHTPADDING', (0, 0), (-1, -1), 0),
                           ('BOX', (0, 6), (0, 6), 0.35, HexColor(0xBDBDBD)),
                           ('BOX', (0, 8), (0, 8), 0.35, HexColor(0xBDBDBD))
                       ]
                       )

        elements.append(t_head)

        doc.build(elements)

        return _buffer

    def __generate_toc(self, queue):
        style_sheet = getSampleStyleSheet()
        style_sheet.add(ParagraphStyle(name="heading_1",
                                       fontName="SourceSansPro-Regular",
                                       fontSize=16,
                                       alignment=TA_LEFT))

        paragraph_header = Paragraph('Contents'.format(self.bold_font), style_sheet["heading_1"])

        data = [[paragraph_header]]

        available_width = self.page_format[0] - (13 * mm * 2)

        t_head = Table(
            data,
            rowHeights=[10],
            colWidths=[available_width],
            style=[
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]
        )

        _buffer = BytesIO()
        can = Canvas(_buffer, pagesize=self.page_format)
        can.setFont("SourceSansPro-Regular", 10)
        can.setFillColor(HexColor(0x000000))
        can.setFillAlpha(1.0)

        width, heigth = self.page_format

        t_head.wrapOn(can, 0, 0)
        t_head.drawOn(can, 13 * mm, heigth - 52)

        vertical_x_limit = int((heigth - (heigth * 0.20)) / 15)
        vertical_x = 1
        num_pages = 1

        for _ in queue:
            vertical_x += 1
            if vertical_x > vertical_x_limit:
                vertical_x = 1
                num_pages += 1

        vertical_x = 1

        top_margin = 60

        current_page_num = num_pages
        for _report in queue:
            section_number = _report.get_section_number()

            indent = len(section_number.split(".")) - 1

            heading_str = "{} {}".format(section_number, _report.name)

            can.drawString(13 * mm + (25 * indent), heigth - (top_margin + (15 * vertical_x)), heading_str)

            heading_str_width = stringWidth(heading_str, "SourceSansPro-Regular", 10)

            dots = "."
            while (25 * indent) + (heading_str_width + 10) + stringWidth(dots, "SourceSansPro-Regular", 10) < \
                    (width - 150):
                dots += "."

            can.setFillAlpha(0.35)
            can.setFillColor(HexColor(0xBDBDBD))
            can.drawString(
                35 + (25 * indent) + (heading_str_width + 10),
                heigth - (top_margin + (15 * vertical_x)),
                dots
            )
            can.setFillAlpha(1.0)
            can.setFillColor(HexColor(0x000000))

            can.drawString(width - 75, heigth - (top_margin + (15 * vertical_x)),
                           "{}".format(current_page_num + num_pages))

            current_page_num += _report.number_of_pages

            vertical_x += 1
            if vertical_x > vertical_x_limit:
                vertical_x = 1
                can.showPage()
                can.setFont("SourceSansPro-Regular", 10)
                t_head.wrapOn(can, 0, 0)
                t_head.drawOn(can, 13 * mm, heigth - 56)

        can.save()

        return _buffer

    def __add_page_numbers(self, pdf_buffer):
        output = PdfFileWriter()
        existing_pdf = PdfFileReader(pdf_buffer)
        num_pages = existing_pdf.getNumPages()

        _tmp_file = tempfile.NamedTemporaryFile()
        self.logo_buffer.seek(0)
        _tmp_file.write(self.logo_buffer.read())
        _tmp_file.flush()

        for page_number in range(num_pages):
            self.set_progress(int((page_number / float(num_pages)) * 90) + 10)
            page = existing_pdf.getPage(page_number)

            if page_number == 0:
                page_num_buffer = BytesIO()
                can = TabbedCanvas(page_num_buffer, pagesize=self.page_format)
                can.setFont("SourceSansPro-Regular", 10)

                tab_spec = (
                    [1.0 * inch, TAB_LEFT | TAB_SPACE],
                    [1.0 * inch, TAB_DECIMAL | TAB_DOT]
                )

                creationdate_str = self.creation_date.strftime(ASSET_DATETIMEFORMAT)
                _user = user.objects.get(idx=self.general_settings["user_idx"])
                _user_str = str(_user)

                str_to_draw_1 = "ServerID\t{}".format(self.cluster_id)
                str_to_draw_2 = "Created\t{} by {}".format(creationdate_str, _user_str)

                can.draw_tabbed_string(25 * mm, 20 * mm, tab_spec, str_to_draw_1)
                can.draw_tabbed_string(25 * mm, 15 * mm, tab_spec, str_to_draw_2)

                can.save()
                page_num_buffer.seek(0)
                page_num_pdf = PdfFileReader(page_num_buffer)
                page.mergePage(page_num_pdf.getPage(0))

            else:
                page_num_buffer = BytesIO()
                can = Canvas(page_num_buffer, pagesize=self.page_format)
                can.setFont("SourceSansPro-Regular", 9)

                # draw page number
                page_number_str = "Page {} of {}".format(page_number + 1, num_pages)

                page_width, page_heigth = self.page_format

                can.drawImage(_tmp_file.name,
                    page_width - (13 * mm) - self.logo_width,
                    page_heigth - 5 * mm - self.logo_height,
                    self.logo_width,
                    self.logo_height,
                    mask="auto")

                can.setFillColor(HexColor(0xBDBDBD))
                can.setStrokeColor(HexColor(0xBDBDBD))
                can.setLineWidth(0.35)
                can.line(13 * mm, 9.5 * mm, page_width - (13 * mm), 9.5 * mm)
                can.setFillColor(HexColor(0x000000))
                can.setStrokeColor(HexColor(0x000000))

                pagee_number_str_width = stringWidth(page_number_str, "SourceSansPro-Regular", 9)

                page_number_str_draw_point = (page_width / 2.0) - (pagee_number_str_width / 2.0)

                can.drawString(page_number_str_draw_point, 6 * mm, page_number_str)

                # draw info footer string
                creationdate_str = self.creation_date.strftime(ASSET_DATETIMEFORMAT)

                info_str = "Report #{}, {}, {}".format(self.report_id, creationdate_str, self.cluster_name)
                info_str_server = None

                if (stringWidth(info_str, "SourceSansPro-Regular", 9) + 13 * mm) > page_number_str_draw_point:
                    info_str = "Report #{}, {},".format(self.report_id, creationdate_str)
                    info_str_server = "{}".format(self.cluster_name)

                can.drawString(13 * mm, 6 * mm, info_str)
                if info_str_server:
                    can.drawString(13 * mm, 3 * mm, info_str_server)

                can.drawImage(NOCTUA_LOGO_PATH, page_width - (13 * mm) - 80, 4 * mm, 80, 18.6, mask="auto")

                can.save()
                page_num_buffer.seek(0)
                page_num_pdf = PdfFileReader(page_num_buffer)
                page.mergePage(page_num_pdf.getPage(0))

            output.addPage(page)

        _tmp_file.close()

        return output

    def __generate_user_group_overview_report(self, root_report):
        _buffer = BytesIO()
        canvas = Canvas(_buffer, self.page_format)

        data = self._get_data_for_user_group_overview()

        data = sorted(data, key=lambda k: (k['group'], k['login']))
        if data:
            report = GenericReport("Userlist")
            root_report.add_child(report)
            section_number = report.get_section_number()

            rpt = PollyReportsReport(data)
            rpt.groupheaders = [Band([Element((0, 4), (self.bold_font, 10),
                                              getvalue=lambda x: x['group'],
                                              format=lambda x: "Group: {}".format(x)), ],
                                     getvalue=lambda x: x["group"])]

            header_names_left = [("Login", "login", 11.0),
                                 ("UID", "uid", 11.0),
                                 ("First name", "firstname", 11.0),
                                 ("Last name", "lastname", 11.0),
                                 ("Email", "email", 11.0),
                                 ("Groups", "secondary_groups", 11.0),
                                 ("Roles", "roles", 11.0),
                                 ("Allowed Device Groups", "allowed_device_groups", 23.0)]

            header_names_right = []

            self.__config_report_helper("{} Userlist".format(section_number),
                                        header_names_left,
                                        header_names_right, rpt, data)

            rpt.generate(canvas)
            canvas.save()
            report.number_of_pages += rpt.pagenumber
            report.add_buffer_to_report(_buffer)
            self.current_page_num += rpt.pagenumber

    def __generate_user_role_overview_report(self, root_report):
        _buffer = BytesIO()
        canvas = Canvas(_buffer, self.page_format)

        data = self._get_data_for_user_roles_overview()

        data = sorted(data, key=lambda k: (k['name']))
        if data:
            report = GenericReport("Rolelist")
            root_report.add_child(report)
            section_number = report.get_section_number()

            rpt = PollyReportsReport(data)

            rpt.groupheaders = [Band([Element((0, 4), (self.bold_font, 10),
                                              getvalue=lambda x: x['name'][0],
                                              format=lambda x: "Roles starting with: {}".format(x)), ],
                                     getvalue=lambda x: x["name"][0]), ]

            header_names_left = [("Name", "name", 33.0),
                                 ("Description", "description", 33.0),
                                 ("Permissions", "permission", 34.0)]

            header_names_right = []

            self.__config_report_helper("{} Rolelist".format(section_number),
                                        header_names_left,
                                        header_names_right, rpt, data)

            rpt.generate(canvas)
            canvas.save()
            report.number_of_pages += rpt.pagenumber
            report.add_buffer_to_report(_buffer)
            self.current_page_num += rpt.pagenumber

    def generate_report(self):
        if self.general_settings["network_report_overview_module_selected"] or \
                self.general_settings["general_device_overview_module_selected"] or \
                self.general_settings["user_group_overview_module_selected"]:

            report = GenericReport("General Reports")
            report.section_number = 1

            self.reports.append(report)

            if self.general_settings["network_report_overview_module_selected"]:
                self.__generate_network_report(report)

            if self.general_settings["general_device_overview_module_selected"]:
                self.__generate_general_device_overview_report(report)

            if self.general_settings["user_group_overview_module_selected"]:
                self.__generate_user_group_overview_report(report)
                self.__generate_user_role_overview_report(report)

        if self.devices:
            device_report = GenericReport("Device Reports")
            device_report.section_number = 1
            if self.reports:
                device_report.section_number = 2

            self.reports.append(device_report)

            group_device_dict = {}
            for _device in self.devices:
                _group_name = _device.device_group_name()
                if _group_name not in group_device_dict:
                    group_device_dict[_group_name] = []

                group_device_dict[_group_name].append(_device)

            idx = 1
            for _group_name in group_device_dict:
                group_report = GenericReport(_group_name)
                device_report.add_child(group_report)

                for _device in sorted(group_device_dict[_group_name], key=lambda __device: __device.full_name):
                    self.__generate_device_report(_device, self.device_settings[_device.idx], group_report)
                    self.set_progress(int((float(idx) / len(self.devices)) * 10))
                    idx += 1

        self.set_progress(10)

        self.finalize_pdf()

    def finalize_pdf(self):
        output_buffer = BytesIO()
        output_pdf = PdfFileWriter()

        queue = []

        def __append_to_queue(report):
            queue.append(report)
            for child in report.children:
                __append_to_queue(child)

        for _report in self.reports:
            __append_to_queue(_report)

        for _report in queue:
            _report.bookmark = None

            for _buffer in _report.pdf_buffers:
                sub_pdf = PdfFileReader(_buffer)
                [output_pdf.addPage(sub_pdf.getPage(page_num)) for page_num in range(sub_pdf.numPages)]

        # generate toc pages, prepend to pdf
        toc_buffer = self.__generate_toc(queue)
        toc_pdf = PdfFileReader(toc_buffer)
        toc_pdf_page_num = toc_pdf.getNumPages()

        for i in reversed(range(toc_pdf_page_num)):
            output_pdf.insertPage(toc_pdf.getPage(i))

        # generate front page, prepend to pdf
        frontpage_buffer = self.__generate_front_page(output_pdf.getNumPages())
        frontpage_pdf = PdfFileReader(frontpage_buffer)
        frontpage_page_num = frontpage_pdf.getNumPages()

        for i in reversed(range(frontpage_page_num)):
            output_pdf.insertPage(frontpage_pdf.getPage(i))

        number_of_pre_content_sites = frontpage_page_num + toc_pdf_page_num

        # Add page numbers
        output_pdf.write(output_buffer)
        output_pdf = self.__add_page_numbers(output_buffer)
        self.set_progress(100)

        # Generate Bookmarks
        current_page_number = number_of_pre_content_sites

        for _report in queue:
            _report.bookmark = output_pdf.addBookmark("{} {}".format(_report.get_section_number(), _report.name),
                                                      current_page_number,
                                                      parent=_report.root_report.bookmark if _report.root_report
                                                      else None)
            current_page_number += _report.number_of_pages

        output_buffer = BytesIO()
        output_pdf.write(output_buffer)
        self.data = output_buffer.getvalue()

        # create report history entry
        _user = user.objects.get(idx=self.general_settings["user_idx"])

        report_history_obj = ReportHistory.objects.get(idx=self.report_id)

        report_history_obj.created_by_user = _user
        report_history_obj.created_at_time = timezone.make_aware(self.creation_date, timezone.get_current_timezone())
        report_history_obj.number_of_pages = output_pdf.getNumPages()
        report_history_obj.size = len(self.data)
        report_history_obj.type = self.get_report_type()
        report_history_obj.generate_filename()
        report_history_obj.write_data(self.data)
        report_history_obj.progress = -1
        report_history_obj.save()

    def get_report_type(self):
        return "pdf"


class XlsxReportGenerator(ReportGenerator):
    def get_report_type(self):
        return "xlsx"

    def generate_report(self):
        workbooks = []

        if self.general_settings['network_report_overview_module_selected']:
            workbook = self.__generate_network_report()

            workbooks.append((workbook, "Network_Overview"))

        if self.general_settings['general_device_overview_module_selected']:
            workbook = self.__generate_general_device_overview_report()

            workbooks.append((workbook, "Device_Overview"))

        if self.general_settings['user_group_overview_module_selected']:
            workbook = self.__generate_user_group_overview_report()

            workbooks.append((workbook, "User_Group_Overview"))

        idx = 1
        for _device in self.devices:
            workbook = Workbook()
            workbook.remove_sheet(workbook.active)

            self.__generate_device_overview(_device, workbook)
            device_report = DeviceReport(_device, self.device_settings[_device.idx], _device.full_name)

            selected_runs = _select_assetruns_for_device(_device,
                                                         self.general_settings["assetbatch_selection_mode"])

            for ar in selected_runs:
                if not device_report.module_selected(ar):
                    continue
                sheet = workbook.create_sheet()
                _title = AssetType(ar.run_type).name

                # xlsx limited to max 31 chars in title
                if len(_title) > 31:
                    _title = _title[:31]

                sheet.title = _title

                generate_csv_entry_for_assetrun(ar, sheet.append)

            self.set_progress(int(round((float(idx) / len(self.devices)) * 100)))
            idx += 1

            workbooks.append((workbook, _device.full_name))

        from zipfile import ZipFile
        _buffer = BytesIO()
        zipfile = ZipFile(_buffer, "w")

        for workbook, workbook_name in workbooks:
            s = save_virtual_workbook(workbook)

            zipfile.writestr("{}.xlsx".format(workbook_name), s)

        zipfile.close()

        self.data = _buffer.getvalue()

        _user = user.objects.get(idx=self.general_settings["user_idx"])
        report_history_obj = ReportHistory.objects.get(idx=self.report_id)

        report_history_obj.created_by_user = _user
        report_history_obj.created_at_time = timezone.make_aware(self.creation_date, timezone.get_current_timezone())
        report_history_obj.number_of_pages = 0
        report_history_obj.size = len(self.data)
        report_history_obj.type = self.get_report_type()
        report_history_obj.generate_filename()
        report_history_obj.write_data(self.data)
        report_history_obj.progress = -1
        report_history_obj.save()

    def __generate_user_group_overview_report(self):
        workbook = Workbook()
        workbook.remove_sheet(workbook.active)

        # user overview
        sheet = workbook.create_sheet()
        sheet.title = "Userlist"

        header_row = ["Login",
                      "UID",
                      "First Name",
                      "Last Name",
                      "Email",
                      "Groups",
                      "Roles",
                      "Allowed Device Groups"
                      ]

        sheet.append(header_row)

        data = self._get_data_for_user_group_overview()

        for entry in data:
            row = [
                entry['login'],
                entry['uid'],
                entry['firstname'],
                entry['lastname'],
                entry['email'],
                entry['secondary_groups'],
                entry['roles'],
                entry['allowed_device_groups']
            ]

            sheet.append(row)

        # role overview
        sheet = workbook.create_sheet()
        sheet.title = "Rolelist"

        header_row = ["Name",
                      "Description",
                      "Permission",
                      ]

        sheet.append(header_row)

        data = self._get_data_for_user_roles_overview()

        for entry in data:
            row = [
                entry['name'],
                entry['description'],
                entry['permission']
            ]

            sheet.append(row)

        return workbook

    @staticmethod
    def __generate_device_overview(_device, workbook):
        sheet = workbook.create_sheet()

        _title = _device.full_name

        # xlsx limited to max 31 chars in title
        if len(_title) > 31:
            _title = _title[:31]

        sheet.title = _title

        header_row = ["FQDN",
                      "DeviceGroup",
                      "DeviceClass",
                      "ComCapabilities",
                      "IP Info",
                      "SNMP Scheme",
                      "SNMP Info",
                      "Device Categories"
                      ]

        sheet.append(header_row)

        data_rows = {}

        idx = 0
        if idx not in data_rows:
            data_rows[idx] = ['' for _ in range(len(header_row))]

        data_rows[idx][0] = _device.full_name
        data_rows[idx][1] = _device.device_group_name()
        data_rows[idx][2] = _device.device_class.name

        idx = 0
        for com_cap in _device.com_capability_list.all():
            if idx not in data_rows:
                data_rows[idx] = ['' for _ in range(len(header_row))]
            data_rows[idx][3] = com_cap.name
            idx += 1

        idx = 0
        for _ip in _device.all_ips():
            if idx not in data_rows:
                data_rows[idx] = ['' for _ in range(len(header_row))]
            data_rows[idx][4] = str(_ip)
            idx += 1

        idx = 0
        for _snmp_scheme in _device.snmp_schemes.all():
            if idx not in data_rows:
                data_rows[idx] = ['' for _ in range(len(header_row))]
            data_rows[idx][5] = str(_snmp_scheme)
            idx += 1

        idx = 0
        for _snmp_scheme in _device.snmp_schemes.all():
            if idx not in data_rows:
                data_rows[idx] = ['' for _ in range(len(header_row))]
            data_rows[idx][7] = str(_snmp_scheme)
            idx += 1

        for i in range(len(data_rows)):
            sheet.append(data_rows[i])

    def __generate_network_report(self):
        from initat.cluster.backbone.models import network

        workbook = Workbook()
        workbook.remove_sheet(workbook.active)

        networks = network.objects.all()

        if networks:
            sheet = workbook.create_sheet()
            sheet.title = "Network Overview"

            header_row = ["Identifier",
                          "Network",
                          "Netmask",
                          "Broadcast",
                          "Gateway",
                          "GW Priority",
                          "#IPs",
                          "Network Type"]

            sheet.append(header_row)

            for _network in networks:
                row = [
                    _network.identifier,
                    _network.network,
                    _network.netmask,
                    _network.broadcast,
                    _network.gateway,
                    _network.gw_pri,
                    _network.num_ip(),
                    _network.network_type.description
                ]

                sheet.append(row)

            for _network in networks:
                sheet = workbook.create_sheet()
                sheet.title = _network.identifier

                header_row = ["IP",
                              "Net-Device",
                              "Device"]

                sheet.append(header_row)

                data = self._get_data_for_sub_network_report(_network)

                for entry in data:
                    row = [
                        entry['ip'],
                        entry['netdevname'],
                        entry['devname']
                    ]

                    sheet.append(row)

        return workbook

    def __generate_general_device_overview_report(self):
        workbook = Workbook()

        data = _generate_hardware_info_data_dict(self.devices, self.general_settings["assetbatch_selection_mode"])

        if data:
            workbook.remove_sheet(workbook.active)

            sheet = workbook.create_sheet("General Device Overview")
            sheet.title = "General Device Overview"

            header_row = ["Name", "Group", "CPU Info", "GPU Info", "Memory Info", "HDD Info", "Partition Info"]

            sheet.append(header_row)

            for _row in data:
                row = [
                    _row['name'],
                    _row['group'],
                    _row['cpu'],
                    _row['gpu'],
                    _row['memory'],
                    _row['hdd'],
                    _row['partition']
                ]

                sheet.append(row)

        return workbook

########################################################################################################################
# Views
########################################################################################################################


class GetProgress(View):
    @method_decorator(login_required)
    def post(self, request):
        report_id = int(request.POST["id"])

        report_history_object = ReportHistory.objects.filter(idx=report_id)

        progress = 0

        if report_history_object:
            report_history_object = report_history_object[0]

            progress = report_history_object.progress

        return HttpResponse(
            json.dumps(
                {
                    'progress': progress
                }
            )
        )


class GetReportData(View):
    @method_decorator(login_required)
    def post(self, request):
        data_b64 = ""
        report_type = "unknown"
        report_id = None

        if "report_id" in request.POST:
            report_id = int(request.POST["report_id"])

            report_history = ReportHistory.objects.get(idx=report_id)
            report_type = report_history.type
            data = report_history.get_data()
            data_b64 = base64.b64encode(data)

        return HttpResponse(
            json.dumps(
                {
                    report_type: data_b64,
                    "report_id": report_id
                }
            )
        )

    @method_decorator(login_required)
    def get(self, request):
        report_id = int(request.GET["report_id"])

        report_history = ReportHistory.objects.get(idx=report_id)
        report_type = report_history.type
        data = report_history.get_data()
        data_b64 = base64.b64encode(data)

        return HttpResponse(
            json.dumps(
                {
                    report_type: data_b64,
                }
            )
        )


class GenerateReportPdf(View):
    @method_decorator(login_required)
    def post(self, request):
        pk_settings, _devices = _init_report_settings(request)

        if 'HOSTNAME' in request.META:
            pk_settings[-1]['hostname'] = request.META['HOSTNAME']
        else:
            pk_settings[-1]['hostname'] = "unknown"

        queue = Queue.Queue()

        Thread(target=_generate_report, args=[pk_settings, _devices, queue, "pdf"]).start()

        report_id = queue.get()

        return HttpResponse(
            json.dumps(
                {
                    'report_id': report_id
                }
            )
        )


class UploadReportGfx(View):
    @method_decorator(xml_wrapper)
    def post(self, request):
        _file = request.FILES[request.FILES.keys()[0]]
        if _file.content_type in ["image/png", "image/jpeg"]:
            system_device = None
            for _device in device.objects.all():
                if _device.is_cluster_device_group():
                    system_device = _device
                    break

            if system_device:
                report_logo_tmp = system_device.device_variable_set.filter(name="__REPORT_LOGO__")
                if report_logo_tmp:
                    report_logo_tmp = report_logo_tmp[0]

                else:
                    report_logo_tmp = device_variable.objects.create(device=system_device, is_public=False,
                                                                     name="__REPORT_LOGO__",
                                                                     inherit=False,
                                                                     protected=True,
                                                                     var_type="b")
                report_logo_tmp.val_blob = base64.b64encode(_file.read())
                report_logo_tmp.save()


class GetReportGfx(View):
    @method_decorator(xml_wrapper)
    def post(self, request):
        id(request)
        val_blob = ""
        system_device = None
        for _device in device.objects.all():
            if _device.is_cluster_device_group():
                system_device = _device
                break

        if system_device:
            report_logo_tmp = system_device.device_variable_set.filter(name="__REPORT_LOGO__")

            if report_logo_tmp:
                val_blob = report_logo_tmp[0].val_blob

        return HttpResponse(
            json.dumps(
                {
                    'gfx': val_blob
                }
            )
        )


class GenerateReportXlsx(View):
    @method_decorator(login_required)
    def post(self, request):
        pk_settings, _devices = _init_report_settings(request)

        queue = Queue.Queue()

        Thread(target=_generate_report, args=[pk_settings, _devices, queue, "xlsx"]).start()

        report_id = queue.get()

        return HttpResponse(
            json.dumps(
                {
                    'report_id': report_id
                }
            )
        )


class ReportDataAvailable(View):
    @method_decorator(login_required)
    def post(self, request):
        idx_list = None
        for item in request.POST.iterlists():
            key, _list = item
            if key == "idx_list[]":
                idx_list = _list
                break

        assetbatch_selection_mode = int(request.POST['assetbatch_selection_mode'])

        pk_setting_dict = {}

        meta_devices = []

        group_selected_runs = {}

        for idx in idx_list:
            idx = int(idx)
            _device = device.objects.get(idx=idx)

            if _device.is_meta_device:
                meta_devices.append(_device)
                continue

            selected_runs = _select_assetruns_for_device(_device, assetbatch_selection_mode)
            selected_run_info_array = \
                [(ar.run_type, str(ar.run_start_time), ar.asset_batch.idx) for ar in selected_runs]

            if _device.device_group_name() not in group_selected_runs:
                group_selected_runs[_device.device_group_name()] = []

            for run_type in [ar.run_type for ar in selected_runs]:
                if run_type not in group_selected_runs[_device.device_group_name()]:
                    group_selected_runs[_device.device_group_name()].append(run_type)

            pk_setting_dict[idx] = selected_run_info_array

        for _device in meta_devices:
            if _device.device_group_name() in group_selected_runs:
                pk_setting_dict[_device.idx] = group_selected_runs[_device.device_group_name()]

        return HttpResponse(
            json.dumps(
                {
                    'pk_setting_dict': pk_setting_dict,
                }
            )
        )


class ReportHistoryAvailable(View):
    @method_decorator(login_required)
    def post(self, request):
        id(request)
        data = {}
        report_ids = []

        for report_history in ReportHistory.objects.all():
            if not report_history.created_by_user:
                continue
            if not report_history.created_at_time:
                continue

            o = {
                'report_id': str(report_history.idx),
                'created_by_user': str(report_history.created_by_user),
                'created_at_time': str(report_history.created_at_time),
                'number_of_pages': str(report_history.number_of_pages),
                'size': sizeof_fmt(report_history.size),
                'raw_size': report_history.b64_size,
                'type': str(report_history.type),
                'number_of_downloads': str(report_history.number_of_downloads)
            }

            report_ids.append(report_history.idx)
            data[report_history.idx] = o

        return HttpResponse(
            json.dumps(
                {
                    'report_ids': list(reversed(sorted(report_ids))),
                    'report_history': data
                }
            )
        )


class UpdateDownloadCount(View):
    @method_decorator(login_required)
    def post(self, request):
        idx = int(request.POST["idx"])

        report_history = ReportHistory.objects.get(idx=idx)

        report_history.number_of_downloads += 1
        report_history.save()

        return HttpResponse(
            json.dumps(
                {
                    'download_count': report_history.number_of_downloads
                }
            )
        )


########################################################################################################################
# Helper Functions
########################################################################################################################

def _init_report_settings(request):
    settings_dict = {}

    for key in request.POST.iterkeys():
        valuelist = request.POST.getlist(key)
        # look for pk in key

        index = key.split("[")[1][:-1]
        if index not in settings_dict:
            settings_dict[index] = {}

        if key[::-1][:4] == ']kp[':
            settings_dict[index]["pk"] = int(valuelist[0])
        else:
            if valuelist[0] == "true":
                value = True
            elif valuelist[0] == "false":
                value = False
            else:
                value = valuelist[0]

            settings_dict[index][key.split("[")[-1][:-1]] = value

    pk_settings = {}

    for setting_index in settings_dict:
        pk = settings_dict[setting_index]['pk']
        pk_settings[pk] = {}

        for key in settings_dict[setting_index]:
            if key != 'pk':
                pk_settings[pk][key] = settings_dict[setting_index][key]

    _devices = []
    for _device in device.objects.filter(idx__in=[int(pk) for pk in pk_settings.keys()]):
        if not _device.is_meta_device:
            _devices.append(_device)

    return pk_settings, _devices


# asset_batch_selection_mode
# 0 == latest only, 1 == latest fully working, 2 == mixed runs from multiple assetbatches
def _select_assetruns_for_device(_device, asset_batch_selection_mode=0):
    selected_asset_runs = []

    asset_batch_selection_mode = int(asset_batch_selection_mode)

    # search latest assetbatch and generate
    if _device.assetbatch_set.all():
        for assetbatch in reversed(sorted(_device.assetbatch_set.all(), key=lambda ab: ab.idx)):
            if asset_batch_selection_mode == 0:
                for assetrun in assetbatch.assetrun_set.all():
                    if assetrun.has_data():
                        selected_asset_runs.append(assetrun)
                break
            elif asset_batch_selection_mode == 1:
                if assetbatch.num_runs == assetbatch.num_completed and assetbatch.num_runs_error == 0:
                    selected_asset_runs = assetbatch.assetrun_set.all()
            elif asset_batch_selection_mode == 2:
                for assetrun in assetbatch.assetrun_set.all():
                    if assetrun.has_data():
                        if AssetType(assetrun.run_type) not in [AssetType(ar.run_type) for ar in selected_asset_runs]:
                            selected_asset_runs.append(assetrun)

    sorted_runs = {}

    for ar in selected_asset_runs:
        if AssetType(ar.run_type) == AssetType.PACKAGE:
            sorted_runs[0] = ar
        elif AssetType(ar.run_type) == AssetType.LICENSE:
            sorted_runs[1] = ar
        elif AssetType(ar.run_type) == AssetType.UPDATE:
            sorted_runs[2] = ar
        elif AssetType(ar.run_type) == AssetType.PENDING_UPDATE:
            sorted_runs[3] = ar
        elif AssetType(ar.run_type) == AssetType.PROCESS:
            # disabled for now
            pass
            # sorted_runs[4] = ar
        elif AssetType(ar.run_type) == AssetType.PRETTYWINHW or AssetType(ar.run_type) == AssetType.LSHW:
            sorted_runs[5] = ar
        elif AssetType(ar.run_type) == AssetType.DMI:
            sorted_runs[6] = ar
        elif AssetType(ar.run_type) == AssetType.PCI:
            sorted_runs[7] = ar
        elif AssetType(ar.run_type) == AssetType.HARDWARE:
            # disabled for now
            pass
            # sorted_runs[8] = ar

    return [sorted_runs[idx] for idx in sorted_runs]


def generate_csv_entry_for_assetrun(ar, row_writer_func):
    base_header = [
        'Asset Type',
        'Batch Id',
        'Start Time',
        'End Time',
        'Total Run Time',
        'device',
        'status',
        'result'
    ]

    if ar.run_start_time and ar.run_end_time:
        ar_run_time = str((ar.run_end_time - ar.run_start_time).total_seconds())
    else:
        ar_run_time = "N/A"

    base_row = [AssetType(ar.run_type).name,
                str(ar.asset_batch.idx),
                str(ar.run_start_time),
                str(ar.run_end_time),
                ar_run_time,
                str(ar.asset_batch.device.full_name),
                RunStatus(ar.run_status).name,
                RunResult(ar.run_result).name]

    if ar.run_type == AssetType.PACKAGE:
        base_header.extend([
            'package_name',
            'package_version',
            'package_release',
            'package_size',
            'package_install_date',
            'package_type'])

        row_writer_func(base_header)

        for package_version in ar.asset_batch.packages.select_related("asset_package").all():
            row = base_row[:]

            row.append(package_version.asset_package.name)
            row.append(package_version.version)
            row.append(package_version.release)
            row.append(package_version.size)
            row.append(package_version.created)
            row.append(PackageTypeEnum(package_version.asset_package.package_type).name)

            row_writer_func(row)

    elif ar.run_type == AssetType.HARDWARE:
        base_header.extend([
            'hardware_node_type',
            'hardware_depth',
            'hardware_attributes',
            'hardware_info'])

        row_writer_func(base_header)

        for hardware_item in ar.assethardwareentry_set.all():
            row = base_row[:]

            row.append(hardware_item.type)
            row.append(hardware_item.depth)
            row.append(hardware_item.attributes)
            row.append(hardware_item.info_list)

            row_writer_func(row)

    elif ar.run_type == AssetType.LICENSE:
        base_header.extend(
            [
                'license_name',
                'license_key'
            ]
        )

        row_writer_func(base_header)

        for _license in ar.assetlicenseentry_set.all():
            row = base_row[:]

            row.append(_license.name)
            row.append(_license.license_key)

            row_writer_func(row)

    elif ar.run_type == AssetType.UPDATE:
        base_header.extend([
            'update_name',
            'update_version',
            'update_release',
            'update_kb_idx',
            'update_install_date',
            'update_status',
            'update_optional',
            'update_installed'
        ])

        row_writer_func(base_header)

        for update in ar.assetupdateentry_set.all():
            row = base_row[:]

            row.append(update.name)
            row.append(update.version)
            row.append(update.release)
            row.append(update.kb_idx)
            row.append(update.install_date)
            row.append(update.status)
            row.append(update.optional)
            row.append(update.installed)

            row_writer_func(row)

    elif ar.run_type == AssetType.PROCESS:
        base_header.extend([
            'process_name',
            'process_id'
        ])

        row_writer_func(base_header)

        for process in ar.assetprocessentry_set.all():
            row = base_row[:]

            row.append(str(process.name))
            row.append(str(process.pid))

            row_writer_func(row)

    elif ar.run_type == AssetType.PENDING_UPDATE:
        base_header.extend([
            'update_name',
            'update_version',
            'update_release',
            'update_kb_idx',
            'update_install_date',
            'update_status',
            'update_optional',
            'update_installed',
            'update_new_version'
        ])

        row_writer_func(base_header)

        for update in ar.assetupdateentry_set.all():
            row = base_row[:]

            row.append(update.name)
            row.append(update.version)
            row.append(update.release)
            row.append(update.kb_idx)
            row.append(update.install_date)
            row.append(update.status)
            row.append(update.optional)
            row.append(update.installed)
            row.append(update.new_version)

            row_writer_func(row)

    elif ar.run_type == AssetType.PCI:
        base_header.extend([
            'domain',
            'bus',
            'slot',
            'func',
            'position',
            'subclass',
            'vendor',
            'device',
            'revision'
        ])

        row_writer_func(base_header)

        for pci_entry in ar.assetpcientry_set.all():
            row = base_row[:]

            # row.append(str(pci_entry))
            row.append(str(pci_entry.domain))
            row.append(str(pci_entry.bus))
            row.append(str(pci_entry.slot))
            row.append(str(pci_entry.func))
            row.append("{:04x}:{:02x}:{:02x}.{:x}".format(pci_entry.domain, pci_entry.bus,
                                                          pci_entry.slot, pci_entry.func))
            row.append(str(pci_entry.subclassname))
            row.append(str(pci_entry.vendorname))
            row.append(str(pci_entry.devicename))
            row.append(str(pci_entry.revision))

            row_writer_func(row)

    elif ar.run_type == AssetType.DMI:
        base_header.extend([
            'handle',
            'dmi_type',
            'header',
            'key',
            'value'
        ])

        row_writer_func(base_header)

        for dmi_head in ar.assetdmihead_set.all():
            for dmi_handle in dmi_head.assetdmihandle_set.all():
                handle = dmi_handle.handle
                dmi_type = dmi_handle.dmi_type
                header = dmi_handle.header

                for dmi_value in dmi_handle.assetdmivalue_set.all():
                    key = dmi_value.key
                    value = dmi_value.value

                    row = base_row[:]

                    row.append(handle)
                    row.append(dmi_type)
                    row.append(header)
                    row.append(key)
                    row.append(value)

                    row_writer_func(row)

    elif ar.run_type == AssetType.PRETTYWINHW or ar.run_type == AssetType.LSHW:
        base_header.extend([
            'entry'
        ])

        row_writer_func(base_header)

        for cpu in ar.asset_batch.cpus.all():
            row = base_row[:]

            row.append(str(cpu))

            row_writer_func(row)

        for memorymodule in ar.asset_batch.memory_modules.all():
            row = base_row[:]

            row.append(str(memorymodule))

            row_writer_func(row)

        for gpu in ar.asset_batch.gpus.all():
            row = base_row[:]

            row.append(str(gpu))

            row_writer_func(row)

        for hdd in ar.asset_batch.hdds.all():
            row = base_row[:]

            row.append(str(hdd))

            row_writer_func(row)

        for partition in ar.asset_batch.partitions.all():
            row = base_row[:]

            row.append(str(partition))

            row_writer_func(row)

        for display in ar.asset_batch.displays.all():
            row = base_row[:]

            row.append(str(display))

            row_writer_func(row)


def sizeof_fmt(num, suffix='B'):
    if num is None:
        return "N/A"
    for unit in ['', 'Ki', 'Mi', 'Gi', 'Ti', 'Pi', 'Ei', 'Zi']:
        if abs(num) < 1024.0:
            return "%3.1f%s%s" % (num, unit, suffix)
        num /= 1024.0
    return "%.1f%s%s" % (num, 'Yi', suffix)


def _generate_report(pk_settings, _devices, queue, type):
    report_generator = None
    try:
        if type == "pdf":
            report_generator = PDFReportGenerator(pk_settings, _devices)
        else:
            report_generator = XlsxReportGenerator(pk_settings, _devices)

        report_id = -1
        if report_generator:
            report_id = report_generator.report_id
        queue.put(report_id)

        report_generator.generate_report()
    except Exception as e:
        # import traceback, sys
        # print '-'*60
        # traceback.print_exc(file=sys.stdout)
        # print '-'*60
        logger.info("Report Generation failed, error was: {}".format(str(e)))

        if report_generator:
            report_generator.data = ""
            report_generator.set_progress(-1)


def _generate_hardware_info_data_dict(_devices, assetbatch_selection_mode):
    data = []

    for _device in _devices:
        selected_runs = _select_assetruns_for_device(_device, assetbatch_selection_mode)

        cpu_str = "N/A"
        gpu_str = "N/A"
        hdd_str = "N/A"
        partition_str = "N/A"
        memory_str = "N/A"

        for assetrun in selected_runs:
            if AssetType(assetrun.run_type) == AssetType.PRETTYWINHW or AssetType(assetrun.run_type) == AssetType.LSHW:
                for cpu in assetrun.asset_batch.cpus.all():
                    if cpu_str != "N/A":
                        cpu_str += "\n{}".format(str(cpu))
                    else:
                        cpu_str = str(cpu)

                for gpu in assetrun.asset_batch.gpus.all():
                    if gpu_str != "N/A":
                        gpu_str += "\n{}".format(str(gpu))
                    else:
                        gpu_str = str(gpu)

                for hdd in assetrun.asset_batch.hdds.all():
                    if hdd_str != "N/A":
                        hdd_str += "\n{}".format(str(hdd))
                    else:
                        hdd_str = str(hdd)

                for partition in assetrun.asset_batch.partitions.all():
                    if partition_str != "N/A":
                        partition_str += "\n{}".format(str(partition))
                    else:
                        partition_str = str(partition)

                for memory_module in assetrun.asset_batch.memory_modules.all():
                    if memory_str != "N/A":
                        memory_str += "\n{}".format(str(memory_module))
                    else:
                        memory_str = str(memory_module)

        o = {
            'name': _device.full_name,
            'group': _device.device_group_name(),
            'cpu': cpu_str,
            'gpu': gpu_str,
            'hdd': hdd_str,
            'partition': partition_str,
            'memory': memory_str
        }

        data.append(o)

    return data


TAB_LEFT = 0
TAB_CENTER = 1
TAB_RIGHT = 2
TAB_DECIMAL = 3
TAB_SPACE = 0x00
TAB_DASH = 0x10
TAB_DOT = 0x20
TAB_LINE = 0x30


class TabbedCanvas(Canvas):
    @staticmethod
    def get_matching_tab(pos, tabspec):
        for ts in tabspec:
            if ts[0] > pos:
                return ts
        return None

    def draw_tabbed_string(self, x, y, tabspec, parts):
        parts = parts.split('\t')
        fn = self._fontname
        fs = self._fontsize
        spc = self.stringWidth(' ', fn, fs)
        dot = self.stringWidth('.', fn, fs)
        dash = self.stringWidth('-', fn, fs)

        self.drawString(x, y, parts[0])
        pos = self.stringWidth(parts[0], fn, fs)
        for part in parts[1:]:
            pw = self.stringWidth(part, fn, fs)
            delta = 0

            ts = self.get_matching_tab(pos, tabspec)
            if ts:
                delta = ts[0] - pos
                align = ts[1] & 0x0f
                fill = ts[1] & 0xf0

                # Figure out the skip space based on the align type.

                if align == TAB_LEFT:
                    pass
                elif align == TAB_CENTER:
                    delta -= pw/2
                elif align == TAB_RIGHT:
                    delta -= pw
                elif align == TAB_DECIMAL:
                    twoparts = part.split('.', 1)
                    p1w = self.stringWidth(twoparts[0], fn, fs)
                    delta -= p1w

                # If delta is now negative, make the delta equal to one
                # space.  Although not necessarily intuitive, this is exactly
                # how Microsoft Word handles this situation.

                if delta < 0:
                    delta = spc

                fillstr = ''
                if fill == TAB_SPACE:
                    pass
                elif fill == TAB_DOT:
                    dots = int((delta-spc-spc) / dot)
                    fillstr = '.'*dots
                elif fill == TAB_DASH:
                    dots = int((delta-spc-spc) / dash)
                    fillstr = '-'*dots
                elif fill == TAB_LINE:
                    self.line(x+pos+spc, y, x+pos+delta, y)

                if fillstr:
                    self.drawString(x+pos+spc, y, fillstr)

            self.drawString(x+pos+delta, y, part)
            pos += delta + pw
