# -*- coding: utf-8 -*-

from django.conf import settings
from django.core.cache import cache
from django.http import HttpResponse
from django.utils.safestring import mark_safe
from email.mime.text import MIMEText
from email.utils import parseaddr, formataddr
from lxml import etree
from lxml.builder import E
import argparse
import cStringIO
import codecs
import commands
import csv
import datetime
import decimal
import django
import email
import email.header
import email.mime
import hashlib
import logging_tools
import math
import optparse
import os
import process_tools
import random
import resource
import reversion
import smtplib
import time
import xml.dom.minidom
import zmq

# to reduce dependencies (cluster-server for example)
try:
    from openpyxl.writer.excel import save_virtual_workbook
    from openpyxl.workbook import Workbook
except:
    openpyxl = None

DEBUG_FILE = "/tmp/lp_debug"
TEMP_DIR = "/tmp/xlsx_olim"


def generate_md5_key(*args):
    new_key = hashlib.md5() # pylint: disable-msg=E1101
    for cur_arg in args:
        new_key.update(unicode(cur_arg)) # pylint: disable-msg=E1101
    return new_key.hexdigest()

class logging_pool(object):
    idle_pool = []
    created = 0
    logger_dict = {}
    zmq_context = zmq.Context()

    @staticmethod
    def debug(what):
        if False: # not settings.IS_WINDOWS:
            if not os.path.isfile(DEBUG_FILE):
                file(DEBUG_FILE, "w").write("%s\n" % (str(datetime.datetime.now())))
                os.chmod(DEBUG_FILE, 0666)
            file(DEBUG_FILE, "a").write("%s\n" % (what))

    @staticmethod
    def get_logger(name="http", **kwargs):
        log_name = "%s.%s" % (settings.REL_SITE_ROOT, name)
        if "init.at.%s" % (log_name) in logging_pool.logger_dict:
            cur_logger = logging_pool.logger_dict["init.at.%s" % (log_name)]
            cur_logger.usecount += 1
            logging_pool.debug("used logger '%s' (usecount is %d, pid is %d)" % (
                log_name,
                cur_logger.usecount,
                os.getpid()))
        else:
            logging_pool.created += 1
            cur_logger = logging_tools.get_logger("%s" % (log_name),
                                                  ["udp:%s" % (kwargs.get("logging_server", "10.240.2.62") or "10.240.2.62")] if settings.IS_WINDOWS else ["uds:/var/lib/logging-server/py_log"],
                                                  init_logger=True,
                                                  zmq=True,
                                                  context=logging_pool.zmq_context)
            cur_logger.log_command("ignore_process_id")
            if "max_file_size" in kwargs:
                cur_logger.log_command("set_file_size %d" % (kwargs["max_file_size"]))
            cur_logger.usecount = 1
            logging_pool.debug("created logger (#%d, pid %d), name '%s'" % (
                logging_pool.created,
                os.getpid(),
                log_name))
            logging_pool.logger_dict[cur_logger.logger.name] = cur_logger
        return cur_logger

    @staticmethod
    def free_logger(cur_logger):
        f_logger_name = cur_logger.logger.name
        logging_pool.logger_dict[f_logger_name].usecount -= 1
        logging_pool.debug("decreased usecount for '%s' to %d (pid %d)" % (
            f_logger_name,
            logging_pool.logger_dict[f_logger_name].usecount,
            os.getpid()
        ))


def call_command(cmd, **kwargs):
    if settings.IS_WINDOWS:
        pipe = os.popen(cmd, "r")
        text = pipe.read()
        sts = pipe.close()
        if sts is None:
            sts = 0
        if text[-1:] == '\n':
            text = text[:-1]
        return sts, text
    else:
        log_cmd = kwargs.get("log_command", None)
        pure_cmd = cmd.split()[0]
        s_time = time.time()
        if not pure_cmd.startswith("/"):
            cmd_found = False
            for s_path in kwargs.get("search_path",
                                     ["/usr/bin", "/bin",
                                      "/usr/sbin", "/sbin",
                                      "/usr/local/bin"]):
                if os.path.isfile("%s/%s" % (s_path, pure_cmd)):
                    cmd = "%s/%s" % (s_path, cmd)
                    cmd_found = True
                    break
        else:
            cmd_found = True
        if cmd_found:
            if "cwd" in kwargs:
                cmd = "cd %s; %s" % (kwargs["cwd"], cmd)
            if log_cmd:
                log_cmd("calling command '%s'%s" % (cmd,
                                                    " (%s)" % (kwargs["info"]) if "info" in kwargs else ""))
            c_stat, c_out = commands.getstatusoutput(cmd)
            e_time = time.time()
            if log_cmd:
                log_cmd(" ... took %s, gave %d (%s)" % (logging_tools.get_diff_time_str(e_time - s_time),
                                                        c_stat,
                                                        logging_tools.get_plural("line", len(c_out.split("\n")))),
                        logging_tools.LOG_LEVEL_WARN if c_stat else logging_tools.LOG_LEVEL_OK)
                if kwargs.get("show_output", False):
                    for l_num, line in enumerate(c_out.split("\n")):
                        log_cmd("   %03d %s" % (l_num + 1, line))
        else:
            if log_cmd:
                log_cmd("command '%s' not found" % (cmd), logging_tools.LOG_LEVEL_ERROR)
            c_stat, c_out = (256, "command not found")
        return c_stat, c_out


def expand_html_body(body_str, **kwargs):
    # remove content-type lines
    new_lines = []
    for line in body_str.split("\n"):
        if line.startswith("<link") and line.count("stylesheet"):
            hr_parts = [part for part in line.split() if part.startswith("href")]
            if hr_parts:
                rel_path = hr_parts[0].split('"')[1]
                if "media_root" in kwargs and "media_path" in kwargs:
                    if kwargs["media_path"].endswith("/") and not kwargs["media_root"].endswith("/"):
                        abs_path = rel_path.replace(kwargs["media_path"], "%s/" % (kwargs["media_root"]))
                    else:
                        abs_path = rel_path.replace(kwargs["media_path"], kwargs["media_root"])
                    if os.path.exists(abs_path):
                        new_lines.append("<style>")
                        new_lines.extend(file(abs_path, "r").read().split("\n"))
                        new_lines.append("</style>")
        else:
            new_lines.append(line)
    if kwargs.get("remove_content_type_line", False):
        if new_lines[0].lower().startswith("content-type"):
            new_lines.pop(0)
    new_text = "\n".join(new_lines)
    return new_text


class progress_counter(object):
    @staticmethod
    def get_counter(**kwargs):
        if "name" in kwargs:
            pc_values = cache.get("pc_%s" % (kwargs["name"]))
            if pc_values:
                return progress_counter(name=kwargs["name"],
                                        max_value=pc_values[0],
                                        act_value=pc_values[1])
            else:
                return None
        else:
            return progress_counter("".join([chr(random.randint(97, 122)) for unused in range(16)]))

    def __init__(self, name, **kwargs):
        self.name = name
        self.max_value = kwargs.get("max_value", 0)
        self.act_value = kwargs.get("act_value", 0)

    def save(self):
        cache.set("pc_%s" % (self.name), [self.max_value, self.act_value])

    def delete(self):
        cache.delete("pc_%s" % (self.name))


class sidebar_element(object):
    def __init__(self, name, **kwargs):
        self.name = unicode(name)
        self.is_folder = kwargs.get("is_folder", False)
        self.node_id = kwargs.get("node_id", None)
        if self.is_folder:
            self.__nodes = []
        self.__args = dict([(key, kwargs[key]) for key in ["isLazy", "url"] if key in kwargs])

    def add_node(self, node):
        if self.is_folder:
            self.__nodes.append(node)

    def get_attribute_string(self):
        attr_list = []
        if self.is_folder:
            attr_list.append("class='folder'")
        if self.node_id:
            attr_list.append("id=\"%s\"" % (self.node_id))
        if self.__args:
            attr_list.append("data=\"%s\"" % (", ".join(["%s: '%s'" % (key, value) for key, value in self.__args.iteritems()])))
        return " ".join(attr_list)

    def get_html(self):
        return mark_safe(u"<li %s>%s%s</li>" % (self.get_attribute_string(),
                                                self.name,
                                                self.get_children("html")))

    def get_json(self):
        ret_dict = {"title": self.name,
                    "isLazy": 1 if self.__args.get("isLazy", False) else 0,
                    "isFolder": 1 if self.is_folder else 0}
        if "url" in self.__args:
            ret_dict["url"] = self.__args["url"]
        return ret_dict

    def get_children(self, mode):
        if self.is_folder and self.__nodes:
            return u"<ul>%s</ul>" % ("".join([getattr(node, "get_%s" % (mode))() for node in self.__nodes]))
        else:
            return ""


def merge_xml(elements):
    if not isinstance(elements, list):
        elements = [elements]
    header = elements[0].split(">", 1)[0]
    # strip xml-header from mother and element
    return "%s><merger>%s</merger>" % (header,
                                       "".join([element.split(">", 1)[1] for element in elements]))


def build_simple_xml_node(master, key, value):
    act_node = master.createElement(key)
    if isinstance(value, (int, long, decimal.Decimal)):
        act_node.setAttribute("type", "int")
        act_node.setAttribute("value", "%d" % (value))
    elif isinstance(value, str):
        act_node.setAttribute("type", "str")
        act_node.setAttribute("value", value)
    elif isinstance(value, bool):
        act_node.setAttribute("type", "bool")
        act_node.setAttribute("value", "1" if value else "0")
    elif isinstance(value, unicode):
        act_node.setAttribute("type", "str")
        act_node.setAttribute("value", value)
    else:
        print "Unknown type %s (%s)" % (type(value), str(value))
    return act_node


def build_simple_xml(head_name, in_dict):
    act_dom = xml.dom.minidom.Document()
    head_info = act_dom.createElement(head_name)
    # head_info.setAttribute("version", "1.0")
    act_dom.appendChild(head_info)
    for key, value in in_dict.iteritems():
        if isinstance(value, list):
            for sub_value in value:
                head_info.appendChild(build_simple_xml_node(act_dom, key, sub_value))
        else:
            head_info.appendChild(build_simple_xml_node(act_dom, key, value))
    return act_dom.toxml("utf-8")


def modify_light(hsv_specs, diff_hue):
    hue, sat, val = hsv_specs
    sat += max(min(diff_hue, 1.0), 0.)
    return (hue, sat, val)


def scale_rgb(rgb_specs, fact, diff):
    ret_val = tuple([min(max(val * fact + diff * val * val / (255 * 255), 0), 255) for val in rgb_specs])
    return ret_val


def hash_cert_number_search(in_str):
    # generate hash for search
    if in_str.count(":"):
        in_str = in_str.split(":", 1)[1]
    return "".join([act_c for act_c in in_str if act_c.isdigit()])


def hash_cert_number(in_str):
    # code from above, sigh...
    map_dict = {u"Ö": "0",
                u"O": "0",
                u"I": "1"}
    s_hash = ""
    for act_c in in_str.upper():
        act_o = ord(act_c)
        if act_c in map_dict:
            new_c = map_dict[act_c]
        elif (act_o < 48) or (act_o > 57 and act_o < 65) or (act_o > 90 and act_o < 97):
            new_c = ""
        elif act_o > 122:
            # warning
            new_c = ""
        else:
            new_c = act_c
        s_hash = u"%s%s" % (s_hash, new_c)
    return s_hash


class unicode_writer(object):
    """
    A CSV unicode writer. Use with csv module from stdlib.
    """
    def __init__(self, func, dialect=csv.excel, encoding="utf-8", **kwds):
        # Redirect output to a queue
        self.queue = cStringIO.StringIO()
        self.writer = csv.writer(self.queue, dialect=dialect, **kwds)
        self.stream = func
        self.encoder = codecs.getincrementalencoder(encoding)()

    def writerow(self, row):
        self.writer.writerow([(s if s is not None else "").encode("utf-8") for s in row])
        # Fetch UTF-8 output from the queue ...
        data = self.queue.getvalue()
        data = data.decode("utf-8")
        # ... and reencode it into the target encoding
        data = self.encoder.encode(data)
        # write to the target stream
        self.stream.write(data)
        # empty queue
        self.queue.truncate(0)

    def writerows(self, rows):
        for row in rows:
            self.writerow(row)


class init_base_object(object):
    """ Provide logging facilities for objects. """
    def __init__(self, name, **kwargs):
        # base object, provides logging facilities
        self.name = name
        if "logger" in kwargs:
            self.__logger = kwargs["logger"]
        else:
            lp_args = {}
            for copy_arg in ["logging_server", "max_file_size"]:
                if copy_arg in kwargs:
                    lp_args[copy_arg] = kwargs[copy_arg]
            self.__logger = logging_pool.get_logger(name, **lp_args)
        self.options = kwargs.get("options", None)
        self._check_options()
        if settings.IS_WINDOWS:
            # write log spacer
            self.log("-" * 50)
        self.__quiet = kwargs.get("quiet", False)
        self._init_step_logging()
        self._log_start()
        self.__mail_lines = []
        self.__mail_lines_log_level = logging_tools.LOG_LEVEL_OK
        self.__step_time = None
        self.__step_num = None
        self.__step_history = None
        self.__act_step = None
        self.__verbose = None

    def mail_log(self, what, log_level=logging_tools.LOG_LEVEL_OK):
        self.log(what, log_level, mail=True)

    def get_logger(self):
        return self.__logger

    def log(self, what, log_level=logging_tools.LOG_LEVEL_OK, **kwargs):
        if self.__verbose:
            print "[%s] %s" % (logging_tools.get_log_level_str(log_level), what)
        self.__logger.log(log_level, what)
        if kwargs.get("mail", False):
            self.__mail_lines.append("[%4s] %s" % (logging_tools.get_log_level_str(log_level), what))
            self.__mail_lines_log_level = max(self.__mail_lines_log_level, log_level)

    def get_mail_lines(self):
        return self.__mail_lines

    def get_mail_lines_log_level(self):
        return self.__mail_lines_log_level

    def _init_step_logging(self):
        self.__act_step, self.__step_num, self.__step_history, self.__step_time = (None, 0, [], None)

    def _get_step_history(self):
        return self.__step_history

    def start_step(self, what=None, *args, **kwargs):
        unknown_kwargs = [key for key in kwargs.iterkeys() if key not in ["mail", "command", "revision"]]
        if unknown_kwargs:
            self.log("unknown keys in kwargs: %s" % (", ".join(unknown_kwargs)),
                     logging_tools.LOG_LEVEL_ERROR,
                     mail=True)
        log_mail = kwargs.get("mail", False)
        with_revison = kwargs.get("revision", False)
        act_time = time.time()
        if self.__act_step:
            # finish
            run_time = act_time - self.__step_time
            self.log("finished step %d (%s) in %s" % (self.__step_num,
                                                      self.__act_step,
                                                      logging_tools.get_diff_time_str(run_time)),
                     mail=log_mail)
            self.__step_history.append((self.__act_step, run_time))
        self.__act_step = what
        self.__step_time = act_time
        if self.__act_step:
            self.__step_num += 1
            self.log("starting step %d: %s" % (self.__step_num,
                                               self.__act_step),
                     mail=log_mail)
            if "command" in kwargs:
                if with_revison:
                    with reversion.revision:
                        getattr(self, kwargs["command"])(*args)
                else:
                    getattr(self, kwargs["command"])(*args)
        else:
            self._step_info(mail=log_mail)

    def get_step_runtime(self, step_name):
        return ([run_time for name, run_time in self.__step_history if name == step_name] + [0])[0]

    def _step_info(self, **kwargs):
        log_mail = kwargs.get("mail", False)
        self.log("executed %s in %s:" % (logging_tools.get_plural("step", len(self.__step_history)),
                                         logging_tools.get_diff_time_str(sum([hist[1] for hist in self.__step_history]))),
                 mail=log_mail)
        for s_idx, (s_name, s_time) in enumerate(self.__step_history):
            self.log("step %d (%-30s) : %s" % (s_idx + 1,
                                               s_name,
                                               logging_tools.get_diff_time_str(s_time)))

    def close(self):
        self.__logger.log(logging_tools.LOG_LEVEL_OK, "CLOSE")

    def _log_start(self):
        if not self.__quiet:
            self.log("starting")
            self.log("options present: %s" % ("yes" if self.options else "no"))
        if self.options:
            if self.__quiet:
                self.log("options:")
            if isinstance(self.options, argparse.Namespace):
                empty_options = argparse.Namespace()
            else:
                empty_options = optparse.Values()
            set_keys = [key for key in dir(self.options) if key not in dir(empty_options)]
            for key in sorted(set_keys):
                value = getattr(self.options, key)
                self.log("%-20s : %-14s is %s" % (key,
                                                  str(type(value)),
                                                  str(getattr(self.options, key))))

    def __del__(self):
        if self.__logger:
            if logging_pool:
                logging_pool.free_logger(self.__logger)
            del self.__logger

    def _check_options(self):
        self.__verbose = False
        if self.options:
            if self.options.verbose:
                self.__verbose = True


class xml_response(object):
    def __init__(self):
        self.reset()

    # pylint: disable-msg=W0201
    def reset(self):
        self.log_buffer = []
        self.val_dict = {}

    def feed_log_line(self, log_level, log_str):
        self.log_buffer.append((log_level, log_str))

    def __setitem__(self, key, value):
        self.val_dict[key] = value

    def __getitem__(self, key):
        return self.val_dict[key]

    def update(self, in_dict):
        for key, value in in_dict.iteritems():
            self[key] = value

    def is_ok(self):
        if self.log_buffer:
            return True if max([log_lev for log_lev, unused in self.log_buffer]) == logging_tools.LOG_LEVEL_OK else False
        else:
            return True

    def build_response(self):
        num_errors, num_warnings = (
            len([True for log_lev, log_str in self.log_buffer if log_lev == logging_tools.LOG_LEVEL_ERROR]),
            len([True for log_lev, log_str in self.log_buffer if log_lev == logging_tools.LOG_LEVEL_WARN]))
        return E.response(
            E.header(
                E.messages(
                    *[E.message(log_str, **{"log_level": "%d" % (log_lev),
                                            "log_level_str": logging_tools.get_log_level_str(log_lev)}) for log_lev, log_str in self.log_buffer]),
                **{"code": "%d" % (max([log_lev for log_lev, log_str in self.log_buffer] + [logging_tools.LOG_LEVEL_OK])),
                   "errors": "%d" % (num_errors),
                   "warnings": "%d" % (num_warnings),
                   "messages": "%d" % (len(self.log_buffer))}),
            E.values(
                *[E.value(value if type(value) == etree._Element else unicode(value), **{
                    "name": key,
                    "type": {
                        int: "integer",
                        long: "integer",
                        str: "string",
                        unicode: "string",
                        float: "float",
                        etree._Element: "xml"}.get(type(value), "unknown")}) for key, value in self.val_dict.iteritems()]
            )
        )

    def __unicode__(self):
        return etree.tostring(self.build_response(), encoding=unicode)

    def create_response(self):
        return HttpResponse(unicode(self),
                            mimetype="application/xml")


def keyword_check(*kwarg_list):
    def decorator(func):
        def _wrapped_view(*args, **kwargs):
            diff_set = set(kwargs.keys()) - set(kwarg_list)
            if diff_set:
                raise KeyError("Invalid keyword arguments: %s" % (str(diff_set)))
            return func(*args, **kwargs)
        return _wrapped_view
    return decorator


def send_emergency_mail(**kwargs):
    # mainly used for windows
    exc_info = process_tools.exception_info()
    if "log_lines" in kwargs:
        mesg = MIMEText("\n".join(kwargs.get("log_lines", [])))
    else:
        mesg = MIMEText("\n".join(exc_info.log_lines))
    mesg["Subject"] = "Python error"
    mesg["From"] = "python-error@init.at"
    mesg["To"] = "oekotex@init.at"
    srv = smtplib.SMTP()
    srv.connect(settings.MAIL_SERVER, 25)
    srv.sendmail(mesg["From"], mesg["To"].split(","), mesg.as_string())
    srv.close()


def create_email(**kwargs):
    body_type = kwargs.get("Body_type", "html")
    header_cs = "utf-8"
    email.charset.add_charset(header_cs, email.charset.QP, email.charset.QP, header_cs)
    if body_type == "text":
        msg_root = email.mime.text.MIMEText(kwargs["Body"], _charset=header_cs)
    else:
        msg_root = email.MIMEMultipart.MIMEMultipart(header_cs) # pylint: disable-msg=E1101
        msg_root.preamble = "This is a multi-part message in MIME-format."
    if "Subject" in kwargs:
        msg_root["Subject"] = email.header.Header(kwargs["Subject"].encode(header_cs),
                                                  header_cs).encode()
    if "From" in kwargs:
        from_name, from_addr = parseaddr(kwargs["From"])
        msg_root["From"] = formataddr((str(email.header.Header(unicode(from_name), header_cs)), from_addr.encode("ascii")))
    if "To" in kwargs:
        if isinstance(kwargs["To"], list):
            # for to_name, to_addr in [parseaddr(cur_addr) for cur_addr in kwargs["To"]]:
                # msg_root["To"] = formataddr((str(email.header.Header(unicode(to_name), header_cs)), to_addr.encode("ascii")))
            # list with str contains more email_addrs split with , or ; ["addr, addr"]
            for cur_addrs in kwargs["To"]:
                for to_name, to_addr in [parseaddr(cur_addr) for cur_addr in sum([sub_split.split(",") for sub_split in cur_addrs.split(";")], [])]:
                    msg_root["To"] = formataddr((str(email.header.Header(unicode(to_name), header_cs)), to_addr.encode("ascii")))
        elif type(kwargs["To"]) in [type(u""), type("")]:
            for to_name, to_addr in [parseaddr(cur_addr) for cur_addr in sum([sub_split.split(",") for sub_split in kwargs["To"].split(";")], [])]:
                msg_root["To"] = formataddr((str(email.header.Header(unicode(to_name), header_cs)), to_addr.encode("ascii")))
        else:
            raise TypeError("unknown type for To-argument '%s': %s" % (str(kwargs["To"]),
                                                                       type(kwargs["To"])))
    # msg_root.add_header("content-transfer-encoding", "quoted-printable")
    if "Body" in kwargs:
        body_txt = kwargs["Body"]
        if body_type == "html":
            body_txt = expand_html_body(body_txt,
                                        media_path=settings.MEDIA_URL,
                                        media_root=settings.MEDIA_ROOT,
                                        remove_content_type_line=True)
        # for text mails the body is added at the top
        if body_type != "text":
            msg_body = email.MIMEText.MIMEText(body_txt, # pylint: disable-msg=E1101
                                               body_type,
                                               header_cs)
            msg_root.attach(msg_body)
    return msg_root


def _fix_header(email_obj, h_name):
    c_list = email_obj[h_name]
    del email_obj[h_name]
    if isinstance(c_list, list):
        c_list = ", ".join(c_list)
    email_obj[h_name] = c_list


def _decode_email_field(cur_el, field_name):
    src_val = getattr(cur_el, field_name)
    if src_val.startswith("=?"):
        decoded = email.header.decode_header(src_val.decode("ascii"))
        if decoded[0][0] != src_val:
            setattr(cur_el, field_name, decoded[0][0])


def decode_email_header(cur_el):
    for field_name in ["subject"]:
        _decode_email_field(cur_el, field_name)


class stream_tee(object):
    """Intercept a stream.
    Invoke like so:
    sys.stdout = StreamTee(sys.stdout)
    See: grid 109 for notes on older version (StdoutTee).
    """
    def __init__(self, target):
        self.target = target

    def write(self, out_str):
        self.target.write(self.intercept(out_str))

    def intercept(self, out_str):
        """Pass-through -- Overload this."""
        return out_str


def add_month(in_date, diff):
    next_year, next_month, next_day = (in_date.year,
                                       in_date.month + diff,
                                       in_date.day)
    while next_month > 12:
        next_year += 1
        next_month -= 12
    while next_month < 1:
        next_year -= 1
        next_month += 12
    while True:
        try:
            next_date = in_date.replace(year=next_year,
                                        month=next_month,
                                        day=next_day)
        except Exception: # pylint: disable-msg=W0703
            next_day -= 1
        else:
            break
    return next_date


def next_bill_date(in_date, interval):
    next_year, next_month = (in_date.year,
                             in_date.month + interval)
    if next_month > 12:
        next_year += 1
        next_month -= 12
    next_date = in_date.replace(year=next_year,
                                month=next_month,
                                day=28)
    while True:
        try:
            next_date = next_date.replace(day=next_date.day + 1)
        except Exception: # pylint: disable-msg=W0703
            break
    return next_date


class safe_stream_filter(stream_tee):
    # Convert string traffic to to something safe.
    def __init__(self, target):
        stream_tee.__init__(self, target)
        self.encoding = "utf-8"
        self.errors = "replace"
        self.encode_to = "utf-8"

    def intercept(self, out_str):
        return out_str.encode(self.encoding, self.errors)


def mk_last_of_month(dt_datetime):
    d_year = int(dt_datetime.strftime("%Y"))
    d_month = int(int(dt_datetime.strftime("%m")) % 12 + 1)
    d_day = 1
    if d_month == 1:
        d_year += 1
    next_month = datetime.date(d_year, d_month, d_day)
    delta = datetime.timedelta(days=1)
    return next_month - delta


def xlsx_export_response(request, xlsx_dict, name=None):
    from initat.core.alfresco.alfresco import alfresco_handler

    if not name:
        name = "xlsx_export"
    file_name = "%s_export.xlsx" % (name)
    # save_virtual_workbook converts the Workbook-object in a buffer
    # it's heavy time-consuming
    print "save_virtual_workbook, this may take a few minutes..."
    _stime = datetime.datetime.now()
    xlsx_content = save_virtual_workbook(xlsx_dict["xlsx"])
    print "save_virtual_workbook done in %s" % (datetime.datetime.now() - _stime)

    if not os.path.isdir(TEMP_DIR):
        os.mkdir(TEMP_DIR)
    file(os.path.join(TEMP_DIR, file_name), "wb").write(xlsx_content)
    xlsx_file = file(os.path.join(TEMP_DIR, file_name), "rb").read()

    a = alfresco_handler(request.log, directory="Olim")
    testfolder = "test/testfolder"
    print "##", a.get_dir_list(testfolder)
    if a.get_dir_list(testfolder):
        a.store_content("%s/%s" % (testfolder, file_name), xlsx_file, create_new_version_if_exists=True)

        uuid = "WRONG"
        request.xml_response["url"] = django.core.urlresolvers.reverse("xlsx:fetch_content", args=[uuid])
    else:
        request.log()
        a.log("folder \"%s\" not found" % (testfolder), logging_tools.LOG_LEVEL_ERROR)
    return request.xml_response.create_response()


def xlsx_export_create(ammount, col_index="", col_names="", sheetname=None):
    if ammount > 10000000:
        # to many resultx raise Ajax request timeout
        raise OverflowError
    if col_index == "" or col_names == "":
        # we don't want to export all Lucene table fields
        raise ValueError
    wb = Workbook()
    ws = wb.get_active_sheet()
    if sheetname:
        ws.title = sheetname
    col_index = col_index.split(" ")
    col_names = col_names.split(";")
    return {"xlsx": wb, "col_index": col_index, "col_names": col_names}


def xlsx_export_append(xlsx_dict, export_list, start_row=0):
    wb = xlsx_dict["xlsx"]
    ws = wb.get_active_sheet()
    col_index = xlsx_dict["col_index"]
    col_names = xlsx_dict["col_names"]
    r_count = start_row
    if start_row == 0:
        c_count = 0
        for head_col in col_names:
            cell = ws.cell(row=r_count, column=c_count)
            cell.value = head_col
            c_count = c_count + 1
    r_count = r_count + 1
    for raw in export_list:
        c_count = 0
        for field in col_index:
            cell = ws.cell(row=r_count, column=c_count)
            cell_attr = raw
            if isinstance(cell_attr, dict):
                if cell_attr.get(field) is not None:
                    cell.value = cell_attr.get(field)
            else:
                for attr_name in field.replace("__", ".").split("."):
                    cell_attr = getattr(cell_attr, attr_name)
                cell.value = unicode(cell_attr)
            c_count = c_count + 1
        r_count = r_count + 1
    return {"xlsx": wb, "col_index": col_index, "col_names": col_names}


class MemoryProfile(object):
    """
    Collect information on maximum memory usage. Use repeated calls to measure()
    and the max_usage attribute.
    """
    def __init__(self):
        self.max_usage = 0

    def _memory_usage(self):
        """ Return memory usage in kB (according to 'man 2 getrusage'"""
        return resource.getrusage(resource.RUSAGE_SELF).ru_maxrss

    def measure(self):
        mem = self._memory_usage()
        # print mem
        if mem > self.max_usage:
            self.max_usage = mem
            # print "Found new max: %s" % mem


def sql_iterator(queryset, step=2000):
    """
    Iterate over queryset in *step* sized chunks. Set *step* to a callable
    to calculate the step size based on the queryset length.
    """
    length = queryset.count()
    if callable(step):
        step_size = step(length)
    else:
        step_size = step
    steps = int(math.ceil(length / float(step_size)))

    for i in xrange(steps):
        for obj in queryset[i * step_size:(i + 1) * step_size]:
            yield obj
