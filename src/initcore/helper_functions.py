#!/usr/bin/python-init -Otu
# -*- coding: utf-8 -*-

""" initcore helper functions """

import sys
import re
import os
import time
import datetime
from django.conf import settings
from django.http import HttpResponse
#from backend.models import email_log, certificate, text_multi_short, user_variable, document, language, country, user, certificate_type, certificate_kind, certificate_status, product_class, standard, email_log_certificate
from django.db.models import Q
import logging_tools
#import cc_const
import optparse
import argparse
import process_tools
import smtplib
from email.mime.text import MIMEText
import xml.dom.minidom
import pprint
from django.core.cache import cache
import decimal
import codecs
import random
import csv
import cStringIO
import cPickle
import commands
import pickle
import base64
import stat
import email
import email.mime
import email.header
from email.utils import parseaddr, formataddr
import copy
import colorsys
from lxml import etree
from lxml.builder import E
from django.utils.hashcompat import md5_constructor
from django.utils.crypto import constant_time_compare, salted_hmac
import hashlib
from django.utils.encoding import smart_str
import reversion
from django.utils.safestring import mark_safe
from email.utils import parseaddr, formataddr
import django.core.urlresolvers
from initcore.alfresco_tools import alfresco_handler, alfresco_content
if settings.ZMQ_LOGGING:
    import zmq
else:
    zmq = None
    
if not settings.IS_WINDOWS:
    from openpyxl.workbook import Workbook
    from openpyxl.writer.excel import save_virtual_workbook

def generate_md5_key(*args):
    new_key = hashlib.md5()
    for cur_arg in args:
        new_key.update(unicode(cur_arg))
    return new_key.hexdigest()

DEBUG_FILE = "/tmp/lp_debug"
TEMP_DIR = "/tmp/.dtf"

class logging_pool(object):
    idle_pool = []
    created = 0
    logger_dict = {}
    if zmq:
        zmq_context = zmq.Context()
    else:
        zmq_context = None
    @staticmethod
    def debug(what):
        if False:#not settings.IS_WINDOWS:
            if not os.path.isfile(DEBUG_FILE):
                file(DEBUG_FILE, "w").write("%s\n" % (str(datetime.datetime.now())))
                os.chmod(DEBUG_FILE, 0666)
            file(DEBUG_FILE, "a").write("%s\n" % (what))
    @staticmethod
    def get_logger(name="http", **kwargs):
        log_name = "%s.%s" % (settings.REL_SITE_ROOT, name)
        if "init.at.%s" % (log_name) in logging_pool.logger_dict:
            cur_logger = logging_pool.logger_dict["init.at.%s" % (log_name)]
            cur_logger.usecount += 1
            logging_pool.debug("used logger '%s' (usecount is %d, pid is %d)" % (
                log_name,
                cur_logger.usecount,
                os.getpid()))
        else:
            logging_pool.created += 1
            cur_logger = logging_tools.get_logger("%s" % (log_name),
                                                  ["udp:%s" % (kwargs.get("logging_server", "10.240.2.62") or "10.240.2.62")] if settings.IS_WINDOWS else ["uds:/var/lib/logging-server/py_log"],
                                                  init_logger=True,
                                                  zmq=True if zmq else False,
                                                  context=logging_pool.zmq_context)
            cur_logger.log_command("ignore_process_id")
            if "max_file_size" in kwargs:
                cur_logger.log_command("set_file_size %d" % (kwargs["max_file_size"]))
            cur_logger.usecount = 1
            logging_pool.debug("created logger (#%d, pid %d), name '%s'" % (
                logging_pool.created,
                os.getpid(),
                log_name))
            logging_pool.logger_dict[cur_logger.logger.name] = cur_logger
##            
##        if logging_pool.idle_pool:
##            cur_logger = logging_pool.idle_pool.pop(0)
##            cur_logger.logger.name = "init.at.%s" % (log_name)
##            logging_pool.debug("used idle logger, name '%s'" % (log_name))
##        else:
##            logging_pool.created += 1
##            cur_logger = logging_tools.get_logger("%s" % (log_name),
##                                                  ["udp:10.240.4.62"] if settings.IS_WINDOWS else ["uds:/var/lib/logging-server/py_log"],
##                                                  init_logger=True)
##            logging_pool.debug("created logger (#%d), name '%s'" % (logging_pool.created,
##                                                                    log_name))
        return cur_logger
    @staticmethod
    def free_logger(cur_logger):
        f_logger_name = cur_logger.logger.name
        logging_pool.logger_dict[f_logger_name].usecount -= 1
        logging_pool.debug("decreased usecount for '%s' to %d (pid %d)" % (
            f_logger_name,
            logging_pool.logger_dict[f_logger_name].usecount,
            os.getpid()
        ))

def call_command(cmd, **kwargs):
    if settings.IS_WINDOWS:
        pipe = os.popen(cmd, "r")
        text = pipe.read()
        sts = pipe.close()
        if sts is None:
            sts = 0
        if text[-1:] == '\n':
            text = text[:-1]
        return sts, text
    else:
        log_cmd = kwargs.get("log_command", None)
        pure_cmd = cmd.split()[0]
        s_time = time.time()
        if not pure_cmd.startswith("/"):
            cmd_found = False
            for s_path in kwargs.get("search_path",
                                     ["/usr/bin" , "/bin",
                                      "/usr/sbin", "/sbin",
                                      "/usr/local/bin"]):
                if os.path.isfile("%s/%s" % (s_path, pure_cmd)):
                    cmd = "%s/%s" % (s_path, cmd)
                    cmd_found = True
                    break
        else:
            cmd_found = True
        if cmd_found:
            if "cwd" in kwargs:
                cmd = "cd %s; %s" % (kwargs["cwd"], cmd)
            if log_cmd:
                log_cmd("calling command '%s'%s" % (cmd,
                                                    " (%s)" % (kwargs["info"]) if "info" in kwargs else ""))
            c_stat, c_out = commands.getstatusoutput(cmd)
            e_time = time.time()
            if log_cmd:
                log_cmd(" ... took %s, gave %d (%s)" % (logging_tools.get_diff_time_str(e_time - s_time),
                                                        c_stat,
                                                        logging_tools.get_plural("line", len(c_out.split("\n")))),
                        logging_tools.LOG_LEVEL_WARN if c_stat else logging_tools.LOG_LEVEL_OK)
                if kwargs.get("show_output", False):
                    for l_num, line in enumerate(c_out.split("\n")):
                        log_cmd("   %03d %s" % (l_num + 1, line))
        else:
            if log_cmd:
                log_cmd("command '%s' not found" % (cmd), logging_tools.LOG_LEVEL_ERROR)
            c_stat, c_out = (256, "command not found")
        return c_stat, c_out

def expand_html_body(body_str, **kwargs):
    # remove content-type lines
    new_lines = []
    for line in body_str.split("\n"):
        if line.startswith("<link") and line.count("stylesheet"):
            hr_parts = [part for part in line.split() if part.startswith("href")]
            if hr_parts:
                rel_path = hr_parts[0].split('"')[1]
                if "media_root" in kwargs and "media_path" in kwargs:
                    if kwargs["media_path"].endswith("/") and not kwargs["media_root"].endswith("/"):
                        abs_path = rel_path.replace(kwargs["media_path"], "%s/" % (kwargs["media_root"]))
                    else:
                        abs_path = rel_path.replace(kwargs["media_path"], kwargs["media_root"])
                    if os.path.exists(abs_path):
                        new_lines.append("<style>")
                        new_lines.extend(file(abs_path, "r").read().split("\n"))
                        new_lines.append("</style>")
        else:
            new_lines.append(line)
    if kwargs.get("remove_content_type_line", False):
        if new_lines[0].lower().startswith("content-type"):
            new_lines.pop(0)
    new_text = "\n".join(new_lines)
    return new_text
  
class progress_counter(object):
    @staticmethod
    def get_counter(**kwargs):
        if "name" in kwargs:
            pc_values = cache.get("pc_%s" % (kwargs["name"]))
            if pc_values:
                return progress_counter(name=kwargs["name"],
                                        max_value=pc_values[0],
                                        act_value=pc_values[1])
            else:
                return None
        else:
            return progress_counter("".join([chr(random.randint(97, 122)) for x in range(16)]))
    def __init__(self, name, **kwargs):
        self.name = name
        self.max_value = kwargs.get("max_value", 0)
        self.act_value = kwargs.get("act_value", 0)
    def save(self):
        cache.set("pc_%s" % (self.name), [self.max_value, self.act_value])
    def delete(self):
        cache.delete("pc_%s" % (self.name))
    
"""
def remove_temporary_documents(log_com, **kwargs):
    temp_docs = document.objects.filter(Q(temporary=True) & (Q(created=None) | Q(created__lt=datetime.datetime.now() - datetime.timedelta(days=kwargs.get("days", 1)))))
    if len(temp_docs):
        temp_paths = [temp_doc._temp_file_path() for temp_doc in temp_docs]
        log_com("removing %s" % (logging_tools.get_plural("temporary document", len(temp_docs))))
        temp_docs.delete()
    else:
        temp_paths = []
    for full_name in temp_paths:
        if os.path.exists(full_name):
            try:
                os.unlink(full_name)
            except:
                log_com("cannot remove temporary file '%s': %s" % (full_name,
                                                                   process_tools.get_except_info()),
                        logging_tools.LOG_LEVEL_ERROR)
            else:
                log_com("removed temporary file '%s'" % (full_name))
        else:
            log_com("temporary file '%s' does not exist" % (full_name),
                    logging_tools.LOG_LEVEL_WARN)
"""

class sidebar_element(object):
    def __init__(self, name, **kwargs):
        self.name = unicode(name)
        self.is_folder = kwargs.get("is_folder", False)
        self.node_id = kwargs.get("node_id", None)
        if self.is_folder:
            self.__nodes = []
        self.__args = dict([(key, kwargs[key]) for key in ["isLazy", "url"] if key in kwargs])
    def add_node(self, node):
        if self.is_folder:
            self.__nodes.append(node)
    def get_attribute_string(self):
        attr_list = []
        if self.is_folder:
            attr_list.append("class='folder'")
        if self.node_id:
            attr_list.append("id=\"%s\"" % (self.node_id))
        if self.__args:
            attr_list.append("data=\"%s\"" % (", ".join(["%s: '%s'" % (key, value) for key, value in self.__args.iteritems()])))
        return " ".join(attr_list)
    def get_html(self):
        return mark_safe(u"<li %s>%s%s</li>" % (self.get_attribute_string(),
                                                self.name,
                                                self.get_children("html")))
    def get_json(self):
        ret_dict = {"title" : self.name,
                    "isLazy" : 1 if self.__args.get("isLazy", False) else 0,
                    "isFolder" : 1 if self.is_folder else 0}
        if "url" in self.__args:
            ret_dict["url"] = self.__args["url"]
        return ret_dict
    def get_children(self, mode):
        if self.is_folder and self.__nodes:
            return u"<ul>%s</ul>" % ("".join([getattr(node, "get_%s" % (mode))() for node in self.__nodes]))
        else:
            return ""

def merge_xml(elements):
    if type(elements) != type([]):
        elements = [elements]
    header = elements[0].split(">", 1)[0]
    # strip xml-header from mother and element
    return "%s><merger>%s</merger>" % (header,
                                       "".join([element.split(">", 1)[1] for element in elements]))

def build_simple_xml_node(master, key, value):
    act_node = master.createElement(key)
    if type(value) in [type(0), type(decimal.Decimal(0)), type(0L)]:
        act_node.setAttribute("type", "int")
        act_node.setAttribute("value", "%d" % (value))
    elif type(value) == type(""):
        act_node.setAttribute("type", "str")
        act_node.setAttribute("value", value)
    elif type(value) == type(True):
        act_node.setAttribute("type", "bool")
        act_node.setAttribute("value", "1" if value else "0")
    elif type(value) == type(u""):
        act_node.setAttribute("type", "str")
        act_node.setAttribute("value", value)
    else:
        print "Unknown type %s (%s)" % (type(value), str(value))
    return act_node

def build_simple_xml(head_name, in_dict):
    act_dom = xml.dom.minidom.Document()
    head_info = act_dom.createElement(head_name)
    #head_info.setAttribute("version", "1.0")
    act_dom.appendChild(head_info)
    for key, value in in_dict.iteritems():
        if type(value) == type([]):
            for sub_value in value:
                head_info.appendChild(build_simple_xml_node(act_dom, key, sub_value))
        else:
            head_info.appendChild(build_simple_xml_node(act_dom, key, value))
    return act_dom.toxml("utf-8")

def modify_light(hsv_specs, diff_hue):
    hue, sat, val = hsv_specs
    sat += max(min(diff_hue, 1.0), 0.)
    return (hue, sat, val)

def scale_rgb(rgb_specs, fact, diff):
    ret_val = tuple([min(max(val * fact + diff * val * val / (255 * 255), 0), 255) for val in rgb_specs])
    return ret_val

def hash_cert_number_search(in_str):
    # generate hash for search
    if in_str.count(":"):
        in_str = in_str.split(":", 1)[1]
    return "".join([act_c for act_c in in_str if act_c.isdigit()])

def hash_cert_number(in_str):
    # code from above, sigh...
    map_dict = {u"Ö" : "0",
                u"O" : "0",
                u"I" : "1"}
    s_hash = ""
    for act_c in in_str.upper():
        act_o = ord(act_c)
        if act_c in map_dict:
            new_c = map_dict[act_c]
        elif (act_o < 48) or (act_o > 57 and act_o < 65) or (act_o > 90 and act_o < 97):
            new_c = ""
        elif act_o > 122:
            # warning
            new_c = ""
        else:
            new_c = act_c
        s_hash = u"%s%s" % (s_hash, new_c)
    return s_hash

class unicode_writer(object):
    def __init__(self, func, dialect=csv.excel, encoding="utf-8", **kwds):
        # Redirect output to a queue
        self.queue = cStringIO.StringIO()
        self.writer = csv.writer(self.queue, dialect=dialect, **kwds)
        self.stream = func
        self.encoder = codecs.getincrementalencoder(encoding)()
    def writerow(self, row):
        self.writer.writerow([(s if s is not None else "").encode("utf-8") for s in row])
        # Fetch UTF-8 output from the queue ...
        data = self.queue.getvalue()
        data = data.decode("utf-8")
        # ... and reencode it into the target encoding
        data = self.encoder.encode(data)
        # write to the target stream
        self.stream.write(data)
        # empty queue
        self.queue.truncate(0)
    def writerows(self, rows):
        for row in rows:
            self.writerow(row)

def customer_name_hash(name, short=None):
    name_hash = hash_string(name)
    if short is not None:
        short_hash = hash_string(short)
        if name_hash.count(short_hash):
            return name_hash
        else:
            return "%s%s" % (name_hash, short_hash)
    else:
        return name_hash
    
def hash_string(in_str):
    in_str = in_str or ""
    ret_str = ""
    rep_dict = {unichr(199)  : "C",
                unichr(208)  : "D",
                unichr(209)  : "N",
                unichr(223)  : "S",
                unichr(268)  : "C",
                unichr(321)  : "L",
                unichr(323)  : "N",
                unichr(346)  : "S",
                unichr(352)  : "S",
                unichr(377)  : "Z",
                unichr(379)  : "Z",
                unichr(381)  : "Z",
                unichr(8211) : "",
                unichr(8220) : "",
                unichr(8221) : "",
                unichr(8222) : "",
                "0" : "O",
                "1" : "I",
                "!" : "I"}
    for t_char, l_range, u_range, add_list in [("A", 192, 198, [260]),
                                               ("E", 200, 203, [280]),
                                               ("I", 204, 207, []),
                                               ("O", 210, 216, [336, 908]),
                                               ("U", 217, 221, [368])]:
        for idx in range(l_range, u_range + 1) + add_list:
            rep_dict[unichr(idx)] = t_char
    for act_c in in_str.upper():
        try:
            if act_c in ur";:., \/<>[](){}+-_*\"'&%$§":
                new_c = ""
            elif act_c in rep_dict:
                new_c = rep_dict[act_c]
            elif ord(act_c) < 128:
                new_c = act_c
            else:
                new_c = ""
            ret_str = "%s%s" % (ret_str, new_c)
        except:
            pass
    return ret_str

class init_base_object(object):
    def __init__(self, name, **kwargs):
        # base object, provides logging facilities
        self.name = name
        if "logger" in kwargs:
            self.__logger = kwargs["logger"]
        else:
            lp_args = {}
            for copy_arg in ["logging_server", "max_file_size"]:
                if copy_arg in kwargs:
                    lp_args[copy_arg] = kwargs[copy_arg]
            self.__logger = logging_pool.get_logger(name, **lp_args)
        self.options = kwargs.get("options", None)
        self._check_options()
        if settings.IS_WINDOWS:
            # write log spacer
            self.log("-" * 50)
        self.__quiet = kwargs.get("quiet", False)
        self._init_step_logging()
        self._log_start()
        self.__mail_lines = []
        self.__mail_lines_log_level = logging_tools.LOG_LEVEL_OK
    def mail_log(self, what, log_level=logging_tools.LOG_LEVEL_OK):
        self.log(what, log_level, mail=True)
    def get_logger(self):
        return self.__logger
    def log(self, what, log_level=logging_tools.LOG_LEVEL_OK, **kwargs):
        if self.__verbose:
            print "[%s] %s" % (logging_tools.get_log_level_str(log_level), what)
        self.__logger.log(log_level, what)
        if kwargs.get("mail", False):
            self.__mail_lines.append("[%4s] %s" % (logging_tools.get_log_level_str(log_level), what))
            self.__mail_lines_log_level = max(self.__mail_lines_log_level, log_level)
    def get_mail_lines(self):
        return self.__mail_lines
    def get_mail_lines_log_level(self):
        return self.__mail_lines_log_level
    def _init_step_logging(self):
        self.__act_step, self.__step_num, self.__step_history, self.__step_time = (None, 0, [], None)
    def _get_step_history(self):
        return self.__step_history
    def start_step(self, what=None, *args, **kwargs):
        unknown_kwargs = [key for key in kwargs.iterkeys() if key not in ["mail", "command", "revision"]]
        if unknown_kwargs:
            self.log("unknown keys in kwargs: %s" % (", ".join(unknown_kwargs)),
                     logging_tools.LOG_LEVEL_ERROR,
                     mail=True)
        log_mail = kwargs.get("mail", False)
        with_revison = kwargs.get("revision", False)
        act_time = time.time()
        if self.__act_step:
            # finish 
            run_time = act_time - self.__step_time
            self.log("finished step %d (%s) in %s" % (self.__step_num,
                                                      self.__act_step,
                                                      logging_tools.get_diff_time_str(run_time)),
                     mail=log_mail)
            self.__step_history.append((self.__act_step, run_time))
        self.__act_step = what
        self.__step_time = act_time
        if self.__act_step:
            self.__step_num += 1
            self.log("starting step %d: %s" % (self.__step_num,
                                               self.__act_step),
                     mail=log_mail)
            if kwargs.has_key("command"):
                if with_revison:
                    with reversion.revision:
                        getattr(self, kwargs["command"])(*args)
                else:
                    getattr(self, kwargs["command"])(*args)
        else:
            self._step_info(mail=log_mail)
    def get_step_runtime(self, step_name):
        return ([run_time for name, run_time in self.__step_history if name == step_name] + [0])[0]
    def _step_info(self, **kwargs):
        log_mail = kwargs.get("mail", False)
        self.log("executed %s in %s:" % (logging_tools.get_plural("step", len(self.__step_history)),
                                         logging_tools.get_diff_time_str(sum([hist[1] for hist in self.__step_history]))),
                 mail=log_mail)
        for s_idx, (s_name, s_time) in enumerate(self.__step_history):
            self.log("step %d (%-30s) : %s" % (s_idx + 1,
                                               s_name,
                                               logging_tools.get_diff_time_str(s_time)))
    def close(self):
        self.__logger.log(logging_tools.LOG_LEVEL_OK, "CLOSE")
    def _log_start(self):
        if not self.__quiet:
            self.log("starting")
            self.log("options present: %s" % ("yes" if self.options else "no"))
        if self.options:
            if self.__quiet:
                self.log("options:")
            if isinstance(self.options, argparse.Namespace):
                empty_options = argparse.Namespace()
            else:
                empty_options = optparse.Values()
            set_keys = [key for key in dir(self.options) if key not in dir(empty_options)]
            for key in sorted(set_keys):
                value = getattr(self.options, key)
                self.log("%-20s : %-14s is %s" % (key,
                                                 str(type(value)),
                                                 str(getattr(self.options, key))))
    def __del__(self):
        if self.__logger:
            if logging_pool:
                logging_pool.free_logger(self.__logger)
            del self.__logger
    def _check_options(self):
        self.__verbose = False
        if self.options:
            if self.options.verbose:
                self.__verbose = True
        
def parser_options_to_dict(opts):
    empty_options = optparse.Values()
    return dict([(key, getattr(opts, key)) for key in [key for key in dir(opts) if key not in dir(empty_options)]])

class xml_response(object):
    def __init__(self):
        self.reset()
    def reset(self):
        self.log_buffer = []
        self.val_dict = {}
    def feed_log_line(self, log_level, log_str):
        self.log_buffer.append((log_level, log_str))
    def __setitem__(self, key, value):
        self.val_dict[key] = value
    def __getitem__(self, key):
        return self.val_dict[key]
    def update(self, in_dict):
        for key, value in in_dict.iteritems():
            self[key] = value
    def is_ok(self):
        if self.log_buffer:
            return True if max([log_lev for log_lev, log_str in self.log_buffer]) == logging_tools.LOG_LEVEL_OK else False
        else:
            return True
    def build_response(self):
        num_errors, num_warnings = (
            len([True for log_lev, log_str in self.log_buffer if log_lev == logging_tools.LOG_LEVEL_ERROR]),
            len([True for log_lev, log_str in self.log_buffer if log_lev == logging_tools.LOG_LEVEL_WARN]))
        return E.response(
            E.header(
                E.messages(
                    *[E.message(log_str, **{"log_level"     : "%d" % (log_lev),
                                            "log_level_str" : logging_tools.get_log_level_str(log_lev)}) for log_lev, log_str in self.log_buffer]),
                **{"code"     : "%d" % (max([log_lev for log_lev, log_str in self.log_buffer] + [logging_tools.LOG_LEVEL_OK])),
                   "errors"   : "%d" % (num_errors),
                   "warnings" : "%d" % (num_warnings),
                   "messages" : "%d" % (len(self.log_buffer))}),
            E.values(
                *[E.value(value if type(value) == etree._Element else unicode(value), **{
                    "name" : key,
                    "type" : {
                        int            : "integer",
                        long           : "integer",
                        str            : "string",
                        unicode        : "string",
                        float          : "float",
                        etree._Element : "xml"}.get(type(value), "unknown")}) for key, value in self.val_dict.iteritems()]
            )
        )
    def __unicode__(self):
        return etree.tostring(self.build_response(), encoding=unicode)
    def create_response(self):
        return HttpResponse(unicode(self),
                            mimetype="application/xml")
    
def keyword_check(*kwarg_list):
    def decorator(func):
        def _wrapped_view(*args, **kwargs):
            diff_set = set(kwargs.keys()) - set(kwarg_list)
            if diff_set:
                raise KeyError, "Invalid keyword arguments: %s" % (str(diff_set))
            return func(*args, **kwargs)
        return _wrapped_view
    return decorator
    
class keyword_checkd(object):
    def __init__(self, *kwarg_list):
        self.__allowed_kwargs = kwarg_list
    def __call__(self, *args, **kwargs):
        return self._func(*args, **kwargs)

class init_logging(object):
    def __init__(self, func):
        """ decorator for catching exceptions and logging, has to be the first decorator in the decorator queue (above login_required) to catch redirects """
        self.__name__ = func.__name__
        self.__logger = logging_pool.get_logger("http")
        self._func = func
    def log(self, what="", log_level=logging_tools.LOG_LEVEL_OK, **kwargs):
        if kwargs.get("request_vars", False):
            self.log_request_vars(kwargs["request"])
        else:
            for_xml = kwargs.get("xml", False)
            if for_xml:
                self.xml_response.feed_log_line(log_level, what)
            self.__logger.log(log_level, "[%s] %s" % (self.__name__, what))
    def write(self, what):
        self.__stdout_buffer = "%s%s" % (self.__stdout_buffer, what)
        if self.__stdout_buffer.endswith("\n"):
            self.__logger.log(logging_tools.LOG_LEVEL_OK, "[%s stdout] %s" % (self.__name__, self.__stdout_buffer.rstrip()))
            self.__stdout_buffer = ""
        # verbose
        if (not settings.IS_WINDOWS) and settings.DEBUG:
            if not isinstance(self.orig_stdout, logging_tools.log_adapter):
                self.orig_stdout.write(what.encode("utf-8", "replace"))
    def log_request_vars(self, request, log_lines=None):
        if not log_lines:
            log_lines = []
        log_lines.extend(["",
                          "method is %s" % (request.method),
                          ""])
        if request.method == "POST":
            log_lines.append("%s defined" % (logging_tools.get_plural("key", len(request.POST))))
            for key in sorted(request.POST):
                log_lines.append(u"%-20s: %s" % (key, unicode(request.POST[key])))
        for line in log_lines:
            self.log(line, logging_tools.LOG_LEVEL_ERROR)
        if not settings.DEBUG:
            send_emergency_mail(log_lines=log_lines)
    def __call__(self, *args, **kwargs):
        s_time = time.time()
        request = args[0]
        if request.user:
            self.log("calling user is '%s'" % (request.user))
        else:
            self.log("no user defined in request", logging_tools.LOG_LEVEL_WARN)
        if hasattr(request, "init_log_counter"):
            # store previous logger
            request.init_log_counter += 1
            self.__prev_logger = request.log
            self.xml_response = request.xml_response
            request.log("nesting init_logging", logging_tools.LOG_LEVEL_WARN)
        else:
            request.init_log_counter = 1
            self.__prev_logger = None
            self.__prev_xml_response = None
            # modify stdout
            self.orig_stdout = sys.stdout
            # stdout buffer
            sys.stdout = self
            # xml response object
            request.xml_response = xml_response()
            self.xml_response = request.xml_response
        self.__stdout_buffer = ""
        request.log = self.log
        try:
            ret_value = self._func(*args, **kwargs)
        except:
            exc_info = process_tools.exception_info()
            log_lines = exc_info.log_lines
            self.log_request_vars(request, log_lines)
            if request.is_ajax():
                error_str = process_tools.get_except_info()
                ret_value = HttpResponse(build_simple_xml("xml_result", {"err_str" : error_str}),
                                         mimetype="application/xml")
            else:
                raise
        finally:
            if not self.__prev_logger:
                # flush buffer
                if self.__stdout_buffer:
                    self.write("\n")
                # restore stdout
                if request.init_log_counter == 1:
                    sys.stdout = self.orig_stdout
        if request.is_ajax() and ret_value["Content-Type"].lower() not in ["application/xml",
                                                                           "application/json"]:
            #print ret_value["Content-Type"]
            ret_value = HttpResponse(build_simple_xml("xml_result", {"err_str" : "session has expired, please login again"}),
                                     mimetype="application/xml")
        if request.is_ajax() and ret_value["Content-Type"] == "application/json" and "callback" in request.GET:            
            # modify content with callback-function for dynatree calls
            ret_value.content = "%s(%s)" % (request.GET["callback"], ret_value.content)
        if logging_pool:
            logging_pool.free_logger(self.__logger)
        if self.__prev_logger:
            # copy previous logger back to struct
            request.log = self.__prev_logger
            self.__prev_logger = None
        request.init_log_counter -= 1
        e_time = time.time()
        self.log("processed in %s" % (logging_tools.get_diff_time_str(e_time - s_time)))
        return ret_value

def send_emergency_mail(**kwargs):
    # mainly used for windows
    exc_info = process_tools.exception_info()
    if "log_lines" in kwargs:
        mesg = MIMEText("\n".join(kwargs.get("log_lines", [])))
    else:
        mesg = MIMEText("\n".join(exc_info.log_lines))
    mesg["Subject"] = "Python error"
    mesg["From"] = "python-error@init.at"
    mesg["To"] = "oekotex@init.at"
    srv = smtplib.SMTP()
    srv.connect(settings.MAIL_SERVER, 25)
    srv.sendmail(mesg["From"], mesg["To"].split(","), mesg.as_string())
    srv.close()

def create_email(**kwargs):
    body_type = kwargs.get("Body_type", "html")
    header_cs = "utf-8"
    email.charset.add_charset(header_cs, email.charset.QP, email.charset.QP, header_cs)
    if body_type == "text":
        msg_root = email.mime.text.MIMEText(kwargs["Body"], _charset=header_cs)
    else:
        msg_root = email.MIMEMultipart.MIMEMultipart(header_cs)
        msg_root.preamble = "This is a multi-part message in MIME-format."
    if "Subject" in kwargs:
        msg_root["Subject"] = email.header.Header(kwargs["Subject"].encode(header_cs),
                                                  header_cs).encode()
    if "From" in kwargs:
        from_name, from_addr = parseaddr(kwargs["From"])
        msg_root["From"] = formataddr((str(email.header.Header(unicode(from_name), header_cs)), from_addr.encode("ascii")))
    if "To" in kwargs:
        if type(kwargs["To"]) == type([]):
            #for to_name, to_addr in [parseaddr(cur_addr) for cur_addr in kwargs["To"]]:
                #msg_root["To"] = formataddr((str(email.header.Header(unicode(to_name), header_cs)), to_addr.encode("ascii")))
            #list with str contains more email_addrs split with , or ; ["addr, addr"]
            for cur_addrs in kwargs["To"]:
                for to_name, to_addr in [parseaddr(cur_addr) for cur_addr in sum([sub_split.split(",") for sub_split in cur_addrs.split(";")], [])]:
                    msg_root["To"] = formataddr((str(email.header.Header(unicode(to_name), header_cs)), to_addr.encode("ascii")))
        elif type(kwargs["To"]) in [type(u""), type("")]:
            for to_name, to_addr in [parseaddr(cur_addr) for cur_addr in sum([sub_split.split(",") for sub_split in kwargs["To"].split(";")], [])]:
                msg_root["To"] = formataddr((str(email.header.Header(unicode(to_name), header_cs)), to_addr.encode("ascii")))
        else:
            raise TypeError, "unknown type for To-argument '%s': %s" % (str(kwargs["To"]),
                                                                        type(kwargs["To"]))
    #msg_root.add_header("content-transfer-encoding", "quoted-printable")
    if "Body" in kwargs:
        body_txt = kwargs["Body"]
        if body_type == "html":
            body_txt = expand_html_body(body_txt,
                                        media_path=settings.MEDIA_URL,
                                        media_root=settings.MEDIA_ROOT,
                                        remove_content_type_line=True)
        #for text mails the body is added at the top
        if body_type != "text":
            msg_body = email.MIMEText.MIMEText(body_txt,
                                               body_type,
                                               header_cs)
            msg_root.attach(msg_body)
    return msg_root

def _fix_header(email_obj, h_name):
    c_list = email_obj[h_name]
    del email_obj[h_name]
    if type(c_list) == type([]):
        c_list = ", ".join(c_list)
    email_obj[h_name] = c_list
    
"""
@keyword_check("alf_handler", "snapshot_info", "index_list", "email_server", "email_server_port", "snapshot_pk")
def send_email(log_com, from_name, to_name, user, cur_mail, **kwargs):
    alf_handler = kwargs.get("alf_handler", None)
    bcc_addr = ["email_log@oekotex.com"]
    cur_mail["Bcc"] = bcc_addr
    smtp = smtplib.SMTP()
    srv_name, srv_port = (kwargs.get("email_server", settings.MAIL_SERVER),
                          kwargs.get("email_server_port", 25))
    # copy to-addresses and append bcc if needed
    to_addr = copy.deepcopy(cur_mail.get_all("To", []))
    if cur_mail["Bcc"]:
        if type(to_addr) == type([]):
            pass
        else:
            to_addr = to_addr.split(", ")
        to_addr.extend(bcc_addr)
    _fix_header(cur_mail, "To")
    _fix_header(cur_mail, "Bcc")
    email_msg = cur_mail.as_string()
    send_lines = ["sending email via %s:%d (%d bytes)" % (srv_name,
                                                          srv_port,
                                                          len(email_msg)),
                  "from is '%s', to is ['%s', '%s']" % (unicode(from_name),
                                                        unicode(to_name),
                                                        unicode(to_addr))]
    smtp.connect(srv_name, srv_port)
    smtp.sendmail(cur_mail["From"], to_addr, email_msg)
    smtp.close()
    send_stat = 0
    for send_line in send_lines:
        log_com(send_line, logging_tools.LOG_LEVEL_ERROR if send_stat else logging_tools.LOG_LEVEL_OK)
    # snapshots the email and sends it
    snapshot_email(log_com,
                   from_name,
                   to_name,
                   user,
                   cur_mail,
                   alf_handler=alf_handler,
                   snapshot_info=kwargs.get("snapshot_info", ""),
                   index_list=kwargs.get("index_list", []),
                   snapshot_pk=kwargs.get("snapshot_pk", 0))
##        snap_obj = email_log()
##        snap_obj.email_from_name = from_name
##        snap_obj.email_to_name   = to_name
##        snap_obj.email_from = ", ".join(cur_mail["From"]) if type(cur_mail["From"]) is type([]) else cur_mail["From"]
##        snap_obj.email_to   = ", ".join(cur_mail["To"]) if type(cur_mail["To"]) is type([]) else cur_mail["To"]
##        snap_obj.subject    = str(cur_mail["Subject"])[:63]
##        snap_obj.body       = email_msg
##        for key, value in kwargs.iteritems():
##            setattr(snap_obj, key, value)
##        snap_obj.user = user
##        snap_obj.save()
        #print snap_obj.get_alfresco_dir()
    return send_stat, send_lines
"""

"""
@keyword_check("alf_handler", "snapshot_info", "index_list", "snapshot_pk")
def snapshot_email(log_com, from_name, to_name, user, cur_mail, **kwargs):
    alf_handler = kwargs.get("alf_handler", None)
    snapshot_pk = kwargs.get("snapshot_pk")
    if snapshot_pk:
        snap_obj = email_log.objects.get(Q(pk=snapshot_pk))
    else:
        snap_obj = email_log()
    if user is not None:
        snap_obj.user = user
    snap_obj.email_from_name = from_name
    snap_obj.email_to_name = to_name
    email_msg = cur_mail.as_string()
    log_com("message length for %s snapshot: %d" % ("new" if snapshot_pk else "present",
                                                    len(email_msg)))
    snap_obj.email_from = ", ".join(cur_mail["From"]) if type(cur_mail["From"]) is type([]) else cur_mail["From"]
    snap_obj.email_to   = ", ".join(cur_mail["To"]) if type(cur_mail["To"]) is type([]) else cur_mail["To"]
    snap_obj.subject    = str(cur_mail["Subject"])[:63]
    if alf_handler:
        snap_obj.body = ""
    else:
        # deprecated
        snap_obj.body = email_msg
    snap_obj.save()
    index_list = kwargs.get("index_list", [])
    if index_list:
        log_com("    index_list has %s: %s" % (logging_tools.get_plural("entry", len(index_list)),
                                               ", ".join(sorted(index_list))))
        for index_cn in index_list:
            new_elc = email_log_certificate(email_log=snap_obj,
                                            cert_number=index_cn)
            new_elc.save()
    if alf_handler:
        try:
            cur_dir = alf_handler.create_folder(snap_obj.get_alfresco_dir(),
                                                check_for_existing=True,
                                                recursive=True,
                                                return_path_on_success=True)
            s_info = kwargs.get("snapshot_info", "")
            alf_handler.store_content("%s/mail_%d.raw" % (cur_dir,
                                                          snap_obj.pk),
                                      email_msg,
                                      author=unicode(snap_obj.user),
                                      description=u"raw data of mail, user '%s'%s" % (unicode(snap_obj.user),
                                                                                      ", %s" % (s_info) if s_info else ""),
                                      created=datetime.datetime.now(),
                                      check_for_existing=True,
                                      create_new_version_if_exists=True,
                                      mimetype="text/plain")
            c_node = alf_handler.get_result_node()
            snap_obj.alfresco_uuid    = c_node["node-uuid"]
            snap_obj.alfresco_version = c_node.get("versionLabel", "???")
            log_com("alfresco uuid/version is %s/%s" % (snap_obj.alfresco_uuid,
                                                        snap_obj.alfresco_version))
            snap_obj.save()
        except:
            log_com("error handling alfresco: %s" % (process_tools.get_except_info()), logging_tools.LOG_LEVEL_ERROR)
    return snap_obj
"""

def _decode_email_field(cur_el, field_name):
    src_val = getattr(cur_el, field_name)
    if src_val.startswith("=?"):
        decoded = email.header.decode_header(src_val.decode("ascii"))
        if decoded[0][0] != src_val:
            setattr(cur_el, field_name, decoded[0][0])
    
def decode_email_header(cur_el):
    for field_name in ["subject"]:
        _decode_email_field(cur_el, field_name)

"""
def get_email_text(**args):
    # args ... various options
    # needed, blueprint
    email_draft = args["email_draft"]
    em_language = args.get("language", None)
    fb_language = args.get("fallback_language", None)
    act_certificate = args.get("certificate", None)
    test_email = args.get("test_email", None)
    if args.get("create_from_to", False):
        # not needed for prolongation email
        # OekoTexStuff/script/Email.asp, 175, one of the most shitty codes
        print "get_email_text():965"
        if email_draft.sender_is_info_email:
            print act_certificate.customer
            print act_certificate.customer.institute
        else:
            print "*"
    localized_subject = email_draft.get_localized_subject(em_language) if not test_email else test_email.get("subject")
    localized_body = email_draft.get_localized_body(em_language) if not test_email else test_email.get("body")
    if len(localized_subject) + len(localized_body) < 10:
        print "***", act_certificate
    if not localized_subject.strip() and fb_language:
        localized_subject = email_draft.get_localized_subject(fb_language)
    if not localized_body.strip() and fb_language:
        localized_body = email_draft.get_localized_body(fb_language)
    ret_dict = {"subject"    : expand_building_block(act_certificate, localized_subject),
                "html_email" : em_language.only_html_mail}
    body = expand_building_block(act_certificate, localized_body, language=em_language)
    #ugly but necessary
    body = body.replace("#Zertifikatsnummer#", act_certificate.cert_number_without_institute())
    if em_language.only_html_mail:
        body = body.replace("\n", "<br>")
    ret_dict["body"] = body
    return ret_dict
"""

"""
def _get_based_on_certificates(**kwargs):
    # get child certificates
    multi_certs = certificate.objects.filter(Q(cert_child__parent=kwargs["certificate"].cert_number) & 
                                             Q(valid_to__gt=datetime.date.today() - datetime.timedelta(days=180))).select_related("customer", "customer__country").order_by("customer__name_1", "customer__city", "customer__country__description")
    ret_array = []
    if multi_certs:
        ret_array.extend(kwargs["arg_list"])
        for multi_cert in multi_certs:
            act_line = "%s%s, %s, %s%s" % ("<li>" if format == "html" else "",
                                           multi_cert.customer.name_1,
                                           multi_cert.customer.city,
                                           multi_cert.customer.country.get_localized_description(kwargs["language"]) if "language" in kwargs and kwargs["language"] else multi_cert.customer.country.description,
                                           "</li>" if format == "html" else "")
            if act_line not in ret_array:
                # no duplicates
                ret_array.append(act_line)
    return ("<br>" if kwargs["format"] == "html" else "\n").join(ret_array)
"""

def _get_header(**kwargs):
    if False:
        # old code
        return "institute/%d/cert/%s" % (kwargs["certificate"].customer.institute_id,
                                         kwargs["var_dict"]["Zertifikatskopf"])
    else:
        # new code
        return kwargs["var_dict"]["Zertifikatskopf"]

def _get_footer(**kwargs):
    if False:
        # old code
        return "institute/%d/cert/%s" % (kwargs["certificate"].customer.institute_id,
                                         kwargs["var_dict"]["Zertifikatsfuss"])
    else:
        # new code
        return kwargs["var_dict"]["Zertifikatsfuss"]
    
def is_email_address(in_str):
    if re.match("^.*@.*\..*$", in_str):
        return True
    else:
        return False
    
"""
def fetch_based_on_certificates(cert, **kwargs):
    bo_dict = dict([(bo_cert.parent, None) for bo_cert in cert.cert_child.all()])
    p_certs = certificate.objects.filter(Q(cert_number__in=bo_dict.keys())).select_related(*kwargs.get("select_related", [])).order_by("-valid_to")
    # build dict for check_certificate_validity
    ccv_dict = {}
    for p_cert in p_certs:
        ccv_dict.setdefault(p_cert.cert_number, []).append(p_cert)
    for cert_num, cert_list in ccv_dict.iteritems():
        ccv_dict[cert_num] = check_certificate_validity(cert_list)
    return sorted([(cert_num, cert_dict) for cert_num, cert_dict in ccv_dict.iteritems()])
"""

"""
def check_certificate_validity(cert_spec, **kwargs):
    # cert spec can be
    # - the certifcate number (as string)
    # - the certificate id (as integer)
    # - the certificate (as object)
    # - a certificate list orderd_by '-valid_to'
    # refer: GueltigkeitEinesZertifikats, global_function.asp, 146
    # helper dict
    h_dict = {"status"      : 0,
              "color"       : cc_const.CERT_BASE_UNKNOWN,
              "css_class"   : "STATUS_NO_INFORMATION",
              "valid_to"    : None,
              "cert_id"     : 0,
              "cert_number" : "",
              "customer"    : None,
              "valid"       : True}
    #print "*", h_dict
    if type(cert_spec) == type([]):
        # received a list of certificates, yummy
        cert_status = cert_spec
        if len([True for act_c in cert_status if act_c.most_recent]):
            cert_status = [act_c for act_c in cert_status if act_c.most_recent]
    elif type(cert_spec) == type(0):
        # Zertifikate_ID given
        cert_status = certificate.objects.filter(Q(idx=cert_spec)).order_by("-valid_to")
    else:
        # cert_number given
        # cert_number_lut present ?
        if "cert_number_lut" in kwargs:
            cert_status = [cur_cert for cur_cert in kwargs["cert_number_lut"].get(cert_spec, []) if cur_cert.most_recent == True]
            if not cert_status:
                cert_status = kwargs["cert_number_lut"].get(cert_spec, [])
                # order according to valid_to
                cert_status = [entry[1] for entry in reversed(sorted([(cur_cert.valid_to, cur_cert) for cur_cert in cert_status if cur_cert.valid_to]))]
        else:
            cert_status = certificate.objects.filter(Q(cert_number__iexact=cert_spec) & Q(most_recent=True))
            if not cert_status:
                cert_status = certificate.objects.filter(Q(cert_number__iexact=cert_spec)).order_by("-valid_to")
    # no more case madness
    upd_dict = {1  : ([0]                               , 1, cc_const.CERT_BASE_NOT_VALID_BUT, "STATUS_APPLICATION_RUNNING", True ),
                2  : ([0, 1, 2, 3, 6, 8, 10]            , 2, cc_const.CERT_BASE_NOT_VALID    , "STATUS_APPLICATION_RUNNING", True ),
                3  : ([0, 1, 3, 6, 8, 10]               , 3, cc_const.CERT_BASE_NOT_VALID    , "STATUS_APPLICATION_RUNNING", True ),
                4  : ([0, 1, 2, 3, 4, 6, 8, 9, 10]      , 4, cc_const.CERT_BASE_STORNO       , "STATUS_NOT_VALID"          , True ),
                5  : ([0, 1, 2, 3, 4, 6, 8, 9, 10]      , 5, cc_const.CERT_BASE_VALID        , "STATUS_VALID"              , True ),
                6  : ([0]                               , 6, None                            , "STATUS_NOT_VALID"          , True ),
                7  : ([0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10], 7, cc_const.CERT_BASE_STORNO       , "STATUS_WITHDRAWN"          , False),
                8  : ([0, 6]                            , 8, cc_const.CERT_BASE_NOT_VALID    , "STATUS_NOT_VALID"          , True ),
                9  : ([0, 1, 2, 3, 6, 8, 9, 10]         , 9, cc_const.CERT_BASE_VALID        , "STATUS_VALID"              , True ),
                10 : ([0, 1, 2, 3, 6, 8, 9, 10]         , 9, cc_const.CERT_BASE_VALID        , "STATUS_NOT_VALID"          , True )}
    for act_cert in cert_status:
        prev_status = h_dict["status"]
        # format: status to match, new status, color, css_class, copy_valid_to
        upd_stuff = upd_dict[act_cert.status_id]
        if h_dict["status"] in upd_stuff[0]:
            h_dict["status"]    = upd_stuff[1]
            h_dict["color"]     = upd_stuff[2]
            h_dict["css_class"] = upd_stuff[3]
            if upd_stuff[4]:
                h_dict["valid_to"] = act_cert.valid_to
        if h_dict["status"] != prev_status:
            h_dict["certificate"]  = act_cert
            h_dict["cert_id"]      = act_cert.idx
            h_dict["cert_number"]  = act_cert.cert_number
            if kwargs.get("fetch_customer", True):
                h_dict["customer"] = act_cert.customer
    if h_dict["css_class"] == "STATUS_NOT_VALID":
        if h_dict["valid_to"] > datetime.date.today() - datetime.timedelta(days=60):
            h_dict["css_class"] = "STATUS_NOT_VALID_1"
        elif h_dict["valid_to"] > datetime.date.today() - datetime.timedelta(days=365):
            h_dict["css_class"] = "STATUS_NOT_VALID_3"
        else:
            h_dict["css_class"] = "STATUS_NOT_VALID_4"
        h_dict["valid"] = False
    return h_dict
"""

def format_number(in_float, **args):
    int_part, float_part = ("%.2f" % (in_float)).split(".")
    new_str = ""
    while int_part:
        new_str = "%s%s%s" % (int_part[-3:], "." if new_str else "", new_str)
        int_part = int_part[:-3]
    if args.get("strip_zeros_from_float", False):
        if float_part.endswith("0"):
            float_part = float_part[:-1]
    return "%s,%s" % (new_str, float_part)

def generate_full_cert_number_from_cust(cust, cert_num):
    # combines the customer with the cert_num to "%d:%s"
    inst_pf = "%d:" % (cust.institute.pk)
    if cert_num.startswith(inst_pf):
        full_cn = cert_num
    else:
        full_cn = "%s%s" % (inst_pf, cert_num)
    return full_cn

def generate_full_cert_number_from_inst(inst, cert_num):
    # combines the customer with the cert_num to "%d:%s"
    inst_pf = "%d:" % (inst.pk)
    if cert_num.startswith(inst_pf):
        full_cn = cert_num
    else:
        full_cn = "%s%s" % (inst_pf, cert_num)
    return full_cn

def remove_local_session(sess_entry):
    try:
        os.unlink("%s/%s" % (settings.SESSION_FILE_PATH,
                             sess_entry[-1]))
    except:
        pass
    
def get_local_sessions(**kwargs):
    sess_ids = [entry for entry in os.listdir(settings.SESSION_FILE_PATH) if entry.startswith(settings.SESSION_COOKIE_NAME)]
    session_list = []
    for sess_id in sess_ids:
        pure_id = sess_id[len(settings.SESSION_COOKIE_NAME):]
        try:
            act_entry = file("%s/%s" % (settings.SESSION_FILE_PATH,
                                        sess_id), "rb").read()
        except:
            pass
        else:
            try:
                decoded = base64.decodestring(act_entry)
            except:
                pass
            else:
                file_date = datetime.datetime.fromtimestamp(os.stat("%s/%s" % (settings.SESSION_FILE_PATH,
                                                                               sess_id))[stat.ST_MTIME])
                # new code
                c_hash, pickled = decoded.split(":", 1)
                # _hash from base.py
                expected_hash = salted_hmac("django.contrib.sessionsSessionStore", pickled).hexdigest()
                if constant_time_compare(c_hash, expected_hash):
                    try:
                        sess_dict = pickle.loads(pickled)
                    except:
                        pass
                    else:
                        if kwargs.get("report_dummy_sessions", False):
                            if not sess_dict.has_key("oeko_user"):
                                session_list.append((pure_id,
                                                     file_date,
                                                     sess_dict,
                                                     sess_id))
                        else:
                            if sess_dict.has_key("oeko_user"):
                                session_list.append((pure_id,
                                                     file_date,
                                                     sess_dict,
                                                     sess_id))
                else:
                    # old code
                    p_part, t_part = (decoded[:-32], decoded[-32:])
                    c_part = md5_constructor(p_part + settings.SECRET_KEY).hexdigest() 
                    if c_part == t_part:
                        try:
                            sess_dict = pickle.loads(p_part)
                        except:
                            pass
                        else:
                            if kwargs.get("report_dummy_sessions", False):
                                if not sess_dict.has_key("oeko_user"):
                                    session_list.append((pure_id,
                                                         file_date,
                                                         sess_dict,
                                                         sess_id))
                            else:
                                if sess_dict.has_key("oeko_user"):
                                    session_list.append((pure_id,
                                                         file_date,
                                                         sess_dict,
                                                         sess_id))
    return session_list

"""
def get_certificates(search_cn, **args):
    # returns a list of all certificates with the given cert_num for the given institute
    # sorted descending by the pk
    # search_cn has to be a valid certificate number (INST:NUMBER)
    if search_cn.count("#"):
        # search includes excerpt
        search_cn, search_en = search_cn.split("#")
        db_query = (Q(cert_number__iexact=search_cn) & Q(excerpt__iexact=search_en))
    else:
        # no excerpt
        db_query = Q(cert_number__iexact=search_cn)
    if args.get("only_pk_list", True):
        db_cert = certificate.objects.filter(db_query).order_by("-pk").values_list("pk", flat=True)
    else:
        db_cert = certificate.objects.filter(db_query).order_by("-pk")
    return db_cert
"""

"""
def expand_building_block(act_certificate, text, **kwargs):
    bb_format = kwargs.get("format", "txt")
    #print text
    if act_certificate.customer_id:
        cert_customer = act_certificate.customer
    else:
        cert_customer = None
    var_dict = {"Absender"           : kwargs.get("from_name", ""),
                "Adressblock"        : kwargs.get("address_block", act_certificate.address_block),
                "ArtikelgruppeStamm" : act_certificate.article_group_text,
                "Ausstellungsdatum"  : kwargs.get("date_of_issue" , act_certificate.date_of_issue.strftime("%d.%m.%Y") if act_certificate.date_of_issue else "-+*+- date_of_issue not set -+*+-"),
                "gueltig_ab"         : act_certificate.valid_from.strftime("%d.%m.%Y") if act_certificate.valid_from else "-+*+- valid_from not set -+*+-",
                "gueltig_bis"        : act_certificate.valid_to.strftime("%d.%m.%Y") if act_certificate.valid_to else "-+*+- valid_to not set -+*+-",
                "date"               : datetime.date.today().strftime("%d.%m.%Y"),
                "Produktklasse"      : act_certificate.product_class,
                "Standard"           : act_certificate.standard,
                "Pruefnummer"        : act_certificate.test_number,
                "Zertifikatsnummer"  : act_certificate.cert_number_without_institute()}
    if cert_customer:
        var_dict.update({
            "InstCity" : cert_customer.institute.city,
            "InstName" : cert_customer.institute.short_name,
            "Institut" : cert_customer.institute.short_name})
    else:
        var_dict.update({
            "InstCity" : "cert_customer not set",
            "InstName" : "cert_customer not set",
            "Institut" : "cert_customer not set"})
    if kwargs.has_key("certificate_block"):
        # correct var_dict
        c_block = kwargs["certificate_block"]
        var_dict.update({"Artikelgruppentext"           : c_block.article_group_text,
                         "Ausstellungsort"              : c_block.place_of_issue,
                         "Institutskontrollbezeichnung" : c_block.institute_check_name,
                         "Zertifikatskopf"              : c_block.header,
                         "Zertifikatsfuss"              : c_block.footer,
                         "Produktklassebeschreibung"    : act_certificate.get_localized_product_class(kwargs["language"]),
                         })
    special_dict = {"base_for"        : _get_based_on_certificates,
                    "Zertifikatskopf" : _get_header,
                    "Zertifikatsfuss" : _get_footer}
    if "cb_func" in kwargs:
        special_dict.update(kwargs["cb_func"])
    if "var_dict" in kwargs:
        var_dict.update(kwargs["var_dict"])
    var_regexp = re.compile("(?P<pre_text>.*?)%s(?P<var_text>.*?)%s(?P<post_text>.*)" % (kwargs.get("start_marker", "<%"),
                                                                                         kwargs.get("end_marker", "%>")),
                            re.MULTILINE | re.DOTALL)
    while True:
        re_match = var_regexp.match(text)
        if re_match:
            # get arguments
            var_text = re_match.group("var_text")
            if var_text.count("(") and var_text.count(")"):
                var_text, arguments = var_text.split("(", 1)
                # remove last parentheses
                arguments = ")".join(arguments.split(")")[:-1])
                # and now some parsing ... at least it works
                act_arg, is_string = ("", False)
                arg_list, idx = ([], 0)
                while idx < len(arguments):
                    char = arguments[idx]
                    next_char = arguments[idx + 1] if idx < len(arguments) - 1 else ""
                    if char == '"':
                        is_string = not is_string
                    elif char == "," and not is_string:
                        arg_list.append(act_arg)
                        act_arg = ""
                    elif is_string and char == "\\" and next_char in ['"', "\\"]:
                        act_arg = "%s%s" % (act_arg, next_char)
                        idx += 1
                    else:
                        act_arg = "%s%s" % (act_arg, char)
                    idx += 1
                if act_arg:
                    arg_list.append(act_arg)
            else:
                arg_list = []
            var_name = var_text.strip()
            # check for the complex lookups
            if var_name in special_dict.keys():
                expanded_var_value = special_dict[var_name](certificate=act_certificate,
                                                            format=bb_format,
                                                            var_dict=var_dict,
                                                            arg_list=arg_list,
                                                            language=kwargs.get("language", None))
            else:
                expanded_var_value = var_dict.get(var_name, "-+*+- variable '%s' not defined -+*+-" % (var_name))
            text = u"%s%s%s" % (re_match.group("pre_text"),
                                expanded_var_value,
                                re_match.group("post_text"))
        else:
            break
    return text
"""

class stream_tee(object):
    """Intercept a stream.
    Invoke like so:
    sys.stdout = StreamTee(sys.stdout)
    See: grid 109 for notes on older version (StdoutTee).
    """
    def __init__(self, target):
        self.target = target
    def write(self, out_str):
        self.target.write(self.intercept(out_str))
    def intercept(self, out_str):
        """Pass-through -- Overload this."""
        return out_str

def add_month(in_date, diff):
    next_year, next_month, next_day = (in_date.year,
                                       in_date.month + diff,
                                       in_date.day)
    while next_month > 12:
        next_year += 1
        next_month -= 12
    while next_month < 1:
        next_year -= 1
        next_month += 12
    while True:
        try:
            next_date = in_date.replace(year=next_year,
                                        month=next_month,
                                        day=next_day)
        except:
            next_day -= 1
        else:
            break
    return next_date
    
def next_bill_date(in_date, interval):
    next_year, next_month = (in_date.year,
                             in_date.month + interval)
    if next_month > 12:
        next_year += 1
        next_month -= 12
    next_date = in_date.replace(year=next_year,
                                month=next_month,
                                day=28)
    while True:
        try:
            next_date = next_date.replace(day=next_date.day + 1)
        except:
            break
    return next_date

"""def get_localized_fields(obj_list, lang, src_field, dst_field, **args):
    # set default values
    if type(src_field) == type([]):
        src_field, src_field_2 = src_field
        tms_dict = dict([(tms.field_id, tms.description) for tms in text_multi_short.objects.filter(Q(language=lang) & Q(field__in=[getattr(getattr(obj, src_field), src_field_2) for obj in obj_list]))])
        for obj in obj_list:
            setattr(obj, dst_field, tms_dict.get(getattr(getattr(obj, src_field), src_field_2), args.get("dummy_value", "for language %s not defined" % (lang))))
    else:
        tms_dict = dict([(tms.field_id, tms.description) for tms in text_multi_short.objects.filter(Q(language=lang) & Q(field__in=[getattr(obj, src_field) for obj in obj_list]))])
        for obj in obj_list:
            setattr(obj, dst_field, tms_dict.get(getattr(obj, src_field), args.get("dummy_value", "for language %s not defined" % (lang))))
"""

class safe_stream_filter(stream_tee):
    # Convert string traffic to to something safe.
    def __init__(self, target):
        stream_tee.__init__(self, target)
        self.encoding = "utf-8"
        self.errors = "replace"
        self.encode_to = "utf-8"
    def intercept(self, out_str):
        return out_str.encode(self.encoding, self.errors)

"""
def get_jqgrid_user_params(request, name, grid_num=1):
    result = []
    for f_num in xrange(grid_num):
        form = "%s%d" % (name, f_num)
        params = get_user_variables(request, form)
        if params.get(form):
            result.append(params.get(form))
        else:
            result.append({"rowNum"  : u"25", "columns" : []})
    return " ".join([":".join([res["rowNum"], ",".join(res["columns"])]) for res in result])
"""

def xlsx_export_create(ammount, col_index="", col_names="", sheetname=None):
    if ammount > 10000000:
        # to many resultx raise Ajax request timeout
        raise OverflowError
    if col_index == "" or col_names == "":
        # we don't want to export all Lucene table fields
        raise ValueError
    wb = Workbook()
    ws = wb.get_active_sheet()
    if sheetname:
        ws.title = sheetname
    col_index = col_index.split(" ")
    col_names = col_names.split(";")
    return {"xlsx" : wb, "col_index" : col_index, "col_names" : col_names}

def xlsx_export_append(xlsx_dict, export_list, start_row=0, lucene=False, sess_lang=None):
    wb = xlsx_dict["xlsx"]
    ws = wb.get_active_sheet()
    col_index = xlsx_dict["col_index"]
    col_names = xlsx_dict["col_names"]
    r_count = start_row
    #print ws.get_highest_row()
    gpm_dict = {}
    cert_status_dict = {}
    cert_type_dict = {}
    cert_kind_dict = {}
    cert_class_dict = {}
    cert_standard_dict = {}
    cert_country_dict = {}
    if start_row == 0:
        c_count = 0
        for head_col in col_names:
            cell = ws.cell(row = r_count, column = c_count)
            cell.value = head_col
            c_count = c_count + 1
    r_count = r_count + 1
    for raw in export_list:
        c_count = 0
        for field in col_index:
            cell = ws.cell(row = r_count, column = c_count)
            #if lucene:
            #    if raw.getField(field) == None:
            #        cell.value = u""
            #        if field == "customer__country__description":
            #            _country = country.objects.get(Q(idx=raw.getField(u"country_id").stringValue()))
            #            if not cert_country_dict.has_key(_country.pk):
            #                cert_country_dict[_country.pk] = _country.get_localized_description(sess_lang.pk)
            #            cell.value = cert_country_dict[_country.pk]
            #    else:
            #        if field in ["cert_type"]:
            #            cert_type_dict = _get_multi_short_dict(certificate_type, int(raw.getField(field).stringValue()), "pk", sess_lang, cert_type_dict)
            #            cell.value = cert_type_dict[int(raw.getField(field).stringValue())]
            #        elif field in ["kind"]:
            #            cert_kind_dict = _get_multi_short_dict(certificate_kind, int(raw.getField(field).stringValue()), "pk", sess_lang, cert_kind_dict)
            #            cell.value = cert_kind_dict[int(raw.getField(field).stringValue())]
            #        elif field in ["status"]:
            #            cert_status_dict = _get_multi_short_dict(certificate_status, int(raw.getField(field).stringValue()), "pk", sess_lang, cert_status_dict)
            #            cell.value = cert_status_dict[int(raw.getField(field).stringValue())]
            #        elif field in ["product_class"]:
            #            cert_class_dict[int(raw.getField(field).stringValue())] = unicode(product_class.objects.get(Q(pk=raw.getField(field).stringValue())).description)
            #            cell.value = cert_class_dict[int(raw.getField(field).stringValue())]
            #        elif field in ["standard"]:
            #            cert_standard_dict[int(raw.getField(field).stringValue())] = unicode(standard.objects.get(Q(pk=raw.getField(field).stringValue())))
            #            cell.value = cert_standard_dict[int(raw.getField(field).stringValue())]
            #        else:
            #            cell.value = unicode(raw.getField(field).stringValue())
            #else:
            cell_attr = raw
            if type(cell_attr) == type({}):
                #print field, "|", cell_attr.get(field)
                if cell_attr.get(field) != None:
                    #if field in ["product__sub_group__description", "product__description", "material__description"]:
                    #    if not gpm_dict.has_key(cell_attr.get(field)):
                    #        gpm_dict[cell_attr.get(field)] = unicode(text_multi_short.objects.select_related("language", "field").get(Q(field=cell_attr.get(field)) & Q(language=sess_lang)))
                    #    cell.value = gpm_dict[cell_attr.get(field)]
                    #    #cell.value = unicode(text_multi_short.objects.select_related("language", "field").get(Q(field=unicode(cell_attr.get(field))) & Q(language=sess_lang)))
                    #elif field in ["certificate__cert_type__pk"]:
                    #    cert_type_dict = _get_multi_short_dict(certificate_type, cell_attr.get(field), "pk", sess_lang, cert_type_dict)
                    #    cell.value = cert_type_dict[cell_attr.get(field)]
                    #elif field in ["certificate__kind__pk"]:
                    #    cert_kind_dict = _get_multi_short_dict(certificate_kind, cell_attr.get(field), "pk", sess_lang, cert_kind_dict)
                    #    cell.value = cert_kind_dict[cell_attr.get(field)]
                    #elif field in ["certificate__status__pk"]:
                    #    cert_status_dict = _get_multi_short_dict(certificate_status, cell_attr.get(field), "pk", sess_lang, cert_status_dict)
                    #    cell.value = cert_status_dict[cell_attr.get(field)]
                    #else:
                    #try:
                    cell.value = cell_attr.get(field)
                    #except Exception as e:
                    #    print e
                    #    print type(cell_attr.get(field))
            else:
                for attr_name in field.replace("__", ".").split("."):
                    #if attr_name == "id" and not attr_name in dir(cell_attr):
                    #    attr_name = "pk"
                    #if attr_name == "active" and not attr_name in dir(cell_attr):
                    #    attr_name = "current"
                    #if attr_name == "web" and not attr_name in dir(cell_attr):
                    #    attr_name = "url"
                    #if attr_name == "email" and not attr_name in dir(cell_attr):
                    #    attr_name = "email_1"
                    #if attr_name in ["kind"]:
                    #    cert_kind_dict = _get_multi_short_dict(certificate_kind, unicode(getattr(cell_attr, attr_name)), "description", sess_lang, cert_kind_dict)
                    #    cell_attr = cert_kind_dict[unicode(getattr(cell_attr, attr_name))]
                    #elif attr_name in ["cert_type"]:
                    #    cert_type_dict = _get_multi_short_dict(certificate_type, unicode(getattr(cell_attr, attr_name)), "description", sess_lang, cert_type_dict)
                    #    cell_attr = cert_type_dict[unicode(getattr(cell_attr, attr_name))]
                    #elif attr_name in ["status"]:
                    #    cert_status_dict = _get_multi_short_dict(certificate_status, unicode(getattr(cell_attr, attr_name)), "description", sess_lang, cert_status_dict)
                    #    cell_attr = cert_status_dict[unicode(getattr(cell_attr, attr_name))]
                    #else:
                    cell_attr = getattr(cell_attr, attr_name)
                cell.value = unicode(cell_attr)
            c_count = c_count + 1
        r_count = r_count + 1
    return {"xlsx" : wb, "col_index" : col_index, "col_names" : col_names}

"""
def _get_multi_short_dict(obj, key, col, lang, ret_dict):
    if not ret_dict.has_key(key):
        if col == "description":
            _key = unicode(obj.objects.get(description__contains=key).text_short)
        elif col == "pk":
            _key = unicode(obj.objects.get(pk=key).text_short)
        if _key != None and _key != "":
            ret_dict[key] = unicode(text_multi_short.objects.select_related("language", "field").get(Q(field=_key) & Q(language=lang)))
    return ret_dict
"""

def xlsx_export_response(request, xlsx_dict, name=None):
    s_time = time.time()
    if not name:
        name = "xlsx_export"
    file_name = "%s_export.xlsx" % (name)
    # save_virtual_workbook converts the Workbook-object in a buffer
    # it's heavy time-consuming
    print "save_virtual_workbook, this may take a few minutes..."
    _stime = datetime.datetime.now()
    xlsx_content = save_virtual_workbook(xlsx_dict["xlsx"])
    print "save_virtual_workbook done in %s" % (datetime.datetime.now() - _stime)
    
    if not os.path.isdir(TEMP_DIR):
        os.mkdir(TEMP_DIR)
    file(os.path.join(TEMP_DIR, file_name), "wb").write(xlsx_content)
    xlsx_file = file(os.path.join(TEMP_DIR, file_name), "rb").read()
    
    a = alfresco_handler(request.log, directory="Olim")
    testfolder = "test/testfolder"
    print "##", a.get_dir_list(testfolder)
    if a.get_dir_list(testfolder):
        a.store_content("%s/%s" % (testfolder, file_name), xlsx_file, create_new_version_if_exists=True) 
        c_node = alfresco_content(None, a.get_result())
        request.xml_response["url"] = django.core.urlresolvers.reverse("xlsx:fetch_content", args=[c_node["node-uuid"]])
    else:
        request.log()
        a.log("folder \"%s\" not found" % (testfolder), logging_tools.LOG_LEVEL_ERROR)
    return request.xml_response.create_response()

"""
def get_most_recent_cert(certnr):
    #all cterificates import, order, certificate, prolongation - except excerpts
    cert = certificate.objects.filter(Q(cert_number=smart_str(certnr)) & Q(cert_type__in=(1, 2, 3)) & Q(status__in=(5, 7, 8))).order_by("-valid_to", "-date_of_issue", "-created")[0:1]
    # better return None then False, otherwise pylint gets confused
    return cert[0] if len(cert) else None
"""

def mk_last_of_month(dt_datetime):
    d_year = int(dt_datetime.strftime("%Y"))
    d_month = int(int(dt_datetime.strftime("%m")) % 12 + 1)
    d_day = 1
    if d_month == 1:
        d_year += 1
    next_month = datetime.date(d_year, d_month, d_day)
    delta = datetime.timedelta(days = 1)
    return next_month - delta

def is_leapyear(iyear):
    if iyear % 400 == 0:
        return True
    elif iyear % 100 == 0:
        return False
    elif iyear % 4 == 0:
        return True
    else:
        return False
    
def parse_filename(fname):
    _pat = re.compile("(\ |\/|#)")
    return _pat.sub("_", fname).replace(u"Ö", u"O").replace(u"ö", u"o")

if __name__ == "__main__":
    print "Loadable module, exiting..."
    sys.exit(-1)
